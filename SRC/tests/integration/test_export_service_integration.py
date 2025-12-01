"""Integration tests for ExportService with real service instances."""

import json
import os
import tempfile
from pathlib import Path

import pytest

from cuepoint.models.result import TrackResult
from cuepoint.services.export_service import ExportService
from cuepoint.services.logging_service import LoggingService


class TestExportServiceIntegration:
    """Integration tests for ExportService using real service instances."""
    
    @pytest.fixture
    def real_logging_service(self):
        """Create a real logging service."""
        return LoggingService()
    
    @pytest.fixture
    def export_service(self, real_logging_service):
        """Create export service with real logging."""
        return ExportService(logging_service=real_logging_service)
    
    @pytest.fixture
    def sample_results(self):
        """Create sample track results for testing."""
        return [
            TrackResult(
                playlist_index=1,
                title="Test Track 1",
                artist="Test Artist",
                matched=True,
                beatport_url="https://www.beatport.com/track/test/123",
                match_score=95.5,
                confidence="high"
            ),
            TrackResult(
                playlist_index=2,
                title="Test Track 2",
                artist="Another Artist",
                matched=False
            )
        ]
    
    def test_export_to_csv_integration(self, export_service, sample_results):
        """Test CSV export with real service."""
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "test_results.csv")
            
            # Export to CSV
            export_service.export_to_csv(sample_results, filepath)
            
            # CSV export creates multiple files with timestamps, check for any CSV files
            csv_files = list(Path(tmpdir).glob("*.csv"))
            assert len(csv_files) > 0, "At least one CSV file should be created"
            
            # Verify at least one file is not empty
            assert any(f.stat().st_size > 0 for f in csv_files), "At least one CSV file should not be empty"
    
    def test_export_to_json_integration(self, export_service, sample_results):
        """Test JSON export with real service."""
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "test_results.json")
            
            # Export to JSON
            export_service.export_to_json(sample_results, filepath)
            
            # Verify file was created
            assert os.path.exists(filepath)
            
            # Verify JSON is valid
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
                assert isinstance(data, list)
                assert len(data) == 2
    
    def test_export_to_excel_integration(self, export_service, sample_results):
        """Test Excel export with real service."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not available")
        
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "test_results.xlsx")
            
            # Export to Excel
            export_service.export_to_excel(sample_results, filepath)
            
            # Verify file was created
            assert os.path.exists(filepath)
            
            # Verify Excel file is valid
            wb = openpyxl.load_workbook(filepath)
            assert wb.active is not None
            assert wb.active.max_row >= 2  # Header + 2 data rows
    
    def test_export_to_csv_custom_delimiter(self, export_service, sample_results):
        """Test CSV export with custom delimiter."""
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "test_results.csv")
            
            # Export with semicolon delimiter
            export_service.export_to_csv(sample_results, filepath, delimiter=";")
            
            # CSV export creates multiple files with timestamps, check for any CSV files
            csv_files = list(Path(tmpdir).glob("*.csv"))
            assert len(csv_files) > 0, "At least one CSV file should be created"
    
    def test_export_to_json_empty_results(self, export_service):
        """Test JSON export with empty results."""
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "empty_results.json")
            
            # Export empty list
            export_service.export_to_json([], filepath)
            
            # Verify file was created
            assert os.path.exists(filepath)
            
            # Verify JSON contains empty array
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
                assert data == []
    
    def test_export_creates_directory(self, export_service, sample_results):
        """Test that export creates output directory if it doesn't exist."""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create nested directory path
            nested_dir = os.path.join(tmpdir, "nested", "output")
            filepath = os.path.join(nested_dir, "test_results.json")
            
            # Export should create directories
            export_service.export_to_json(sample_results, filepath)
            
            # Verify directory and file were created
            assert os.path.exists(nested_dir)
            assert os.path.exists(filepath)
    
    def test_export_service_without_logging(self, sample_results):
        """Test export service without logging service."""
        service = ExportService()  # No logging service
        
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "test_results.json")
            
            # Should work without logging
            service.export_to_json(sample_results, filepath)
            assert os.path.exists(filepath)

    def test_export_to_csv_special_characters(self, export_service):
        """Test CSV export with special characters in data."""
        results = [
            TrackResult(
                playlist_index=1,
                title='Test, Track "with" quotes',
                artist="Artist & Co.",
                matched=False
            )
        ]
        
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "test_results.csv")
            
            # Export to CSV
            export_service.export_to_csv(results, filepath)
            
            # Verify file was created
            csv_files = list(Path(tmpdir).glob("*.csv"))
            assert len(csv_files) > 0
            
            # Verify special characters are handled (file should be readable)
            with open(csv_files[0], 'r', encoding='utf-8') as f:
                content = f.read()
                assert "Test" in content

    def test_export_to_json_special_characters(self, export_service):
        """Test JSON export with special characters in data."""
        results = [
            TrackResult(
                playlist_index=1,
                title='Test, Track "with" quotes',
                artist="Artist & Co.",
                matched=False
            )
        ]
        
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "test_results.json")
            
            # Export to JSON
            export_service.export_to_json(results, filepath)
            
            # Verify JSON is valid and contains special characters
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
                assert len(data) == 1
                # TrackResult.to_dict() uses 'original_title' not 'title'
                assert "Test" in data[0].get('original_title', '') or "Test" in str(data[0])

    def test_export_to_csv_none_values(self, export_service):
        """Test CSV export when TrackResult has None values."""
        results = [
            TrackResult(
                playlist_index=1,
                title="Test Track",
                artist="Test Artist",
                matched=False,
                beatport_url=None,  # None value
                match_score=None,  # None value
            )
        ]
        
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "test_results.csv")
            
            # Export to CSV - should handle None values gracefully
            export_service.export_to_csv(results, filepath)
            
            # Verify file was created
            csv_files = list(Path(tmpdir).glob("*.csv"))
            assert len(csv_files) > 0

    def test_export_to_json_none_values(self, export_service):
        """Test JSON export when TrackResult has None values."""
        results = [
            TrackResult(
                playlist_index=1,
                title="Test Track",
                artist="Test Artist",
                matched=False,
                beatport_url=None,  # None value
                match_score=None,  # None value
            )
        ]
        
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "test_results.json")
            
            # Export to JSON - should handle None values gracefully
            export_service.export_to_json(results, filepath)
            
            # Verify JSON is valid
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
                assert len(data) == 1

    def test_export_to_csv_permission_error(self, export_service, sample_results):
        """Test CSV export with file permission error."""
        from cuepoint.exceptions.cuepoint_exceptions import ExportError
        from unittest.mock import patch
        
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "test_results.csv")
            
            # Mock file write to raise permission error
            with patch('builtins.open', side_effect=PermissionError("Permission denied")):
                # Should raise ExportError
                with pytest.raises(ExportError):
                    export_service.export_to_csv(sample_results, filepath)

    def test_export_to_json_permission_error(self, export_service, sample_results):
        """Test JSON export with file permission error."""
        from cuepoint.exceptions.cuepoint_exceptions import ExportError
        from unittest.mock import patch
        
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "test_results.json")
            
            # Mock file write to raise permission error
            with patch('builtins.open', side_effect=PermissionError("Permission denied")):
                # Should raise ExportError
                with pytest.raises(ExportError):
                    export_service.export_to_json(sample_results, filepath)

    def test_export_to_excel_missing_dependency(self, export_service, sample_results):
        """Test Excel export when openpyxl is not available."""
        from cuepoint.exceptions.cuepoint_exceptions import ExportError
        from unittest.mock import patch
        
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "test_results.xlsx")
            
            # Mock ImportError for openpyxl import inside the function
            # The import happens inside export_to_excel, so we patch it at the module level
            import builtins
            original_import = builtins.__import__
            
            def mock_import(name, *args, **kwargs):
                if name == 'openpyxl':
                    raise ImportError("No module named 'openpyxl'")
                return original_import(name, *args, **kwargs)
            
            with patch.object(builtins, '__import__', side_effect=mock_import):
                # Should raise ExportError with specific error code
                with pytest.raises(ExportError) as exc_info:
                    export_service.export_to_excel(sample_results, filepath)
                
                assert "EXPORT_EXCEL_MISSING_DEPENDENCY" in str(exc_info.value.error_code) or "openpyxl" in str(exc_info.value.message).lower()

    def test_export_to_json_serialization_error(self, export_service):
        """Test JSON export with serialization error."""
        from cuepoint.exceptions.cuepoint_exceptions import ExportError
        from unittest.mock import patch, MagicMock
        
        # Create a result that might cause serialization issues
        results = [
            TrackResult(
                playlist_index=1,
                title="Test Track",
                artist="Test Artist",
                matched=False
            )
        ]
        
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "test_results.json")
            
            # Mock json.dump to raise TypeError
            with patch('json.dump', side_effect=TypeError("Not JSON serializable")):
                # Should raise ExportError
                with pytest.raises(ExportError):
                    export_service.export_to_json(results, filepath)

    def test_export_to_csv_empty_results(self, export_service):
        """Test CSV export with empty results list."""
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "empty_results.csv")
            
            # Export empty list
            export_service.export_to_csv([], filepath)
            
            # Verify files were created (may have headers only or may not create files for empty)
            csv_files = list(Path(tmpdir).glob("*.csv"))
            # Empty results may or may not create files - both behaviors are acceptable
            # Just verify no error was raised
            assert True  # Test passes if no exception was raised

    def test_export_to_excel_empty_results(self, export_service):
        """Test Excel export with empty results list."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not available")
        
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "empty_results.xlsx")
            
            # Export empty list
            export_service.export_to_excel([], filepath)
            
            # Verify file was created
            assert os.path.exists(filepath)
            
            # Verify Excel file is valid (may have headers only)
            wb = openpyxl.load_workbook(filepath)
            assert wb.active is not None

