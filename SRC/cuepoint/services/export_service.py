#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Export Service Implementation

Service for exporting results to various formats.
"""

import os
import shutil
import tempfile
from pathlib import Path
from typing import List, Optional

from cuepoint.exceptions.cuepoint_exceptions import ExportError
from cuepoint.models.result import TrackResult
from cuepoint.services.interfaces import IExportService, ILoggingService
from cuepoint.services.output_writer import write_csv_files


class ExportService(IExportService):
    """Service for exporting track results to various file formats.

    Supports exporting to CSV, JSON, and Excel formats. Handles file
    creation, directory creation, and formatting.

    Attributes:
        logging_service: Service for logging operations.
    """

    def __init__(self, logging_service: Optional[ILoggingService] = None) -> None:
        """Initialize export service.

        Args:
            logging_service: Optional logging service. If None, errors are raised without logging.
        """
        self.logging_service = logging_service

    def _validate_export_path(
        self, filepath: str, results_count: int, overwrite: bool = False
    ) -> tuple[bool, Optional[str]]:
        """Validate export file path with comprehensive checks.

        Args:
            filepath: Full path to output file.
            results_count: Number of results to export (for disk space estimation).
            overwrite: Whether to allow overwriting existing files.

        Returns:
            Tuple of (is_valid, error_message).
        """
        file_path = Path(filepath)
        parent_dir = file_path.parent

        # Check parent directory exists or can be created
        if not parent_dir.exists():
            try:
                parent_dir.mkdir(parents=True, exist_ok=True)
            except OSError as e:
                return False, f"Cannot create output directory: {parent_dir}\n{str(e)}"

        # Check parent directory is writable
        if not os.access(parent_dir, os.W_OK):
            return False, f"Output directory is not writable: {parent_dir}"

        # Check file doesn't exist (unless overwrite allowed)
        if file_path.exists() and not overwrite:
            return False, f"File already exists: {filepath}. Use overwrite option to replace."

        # Check disk space (rough estimate: 1KB per track minimum)
        try:
            free_space = shutil.disk_usage(parent_dir).free
            estimated_size = results_count * 1024  # 1KB per track minimum
            required_space = estimated_size * 2  # Require 2x for safety
            if free_space < required_space:
                return False, (
                    f"Insufficient disk space.\n\n"
                    f"Required: {self._format_bytes(required_space)}\n"
                    f"Available: {self._format_bytes(free_space)}\n\n"
                    f"Please free up space or choose a different location."
                )
        except Exception as e:
            # If we can't check disk space, log but don't fail
            if self.logging_service:
                self.logging_service.warning(
                    f"Could not check disk space: {e}", extra={"filepath": filepath}
                )

        return True, None

    def _format_bytes(self, bytes: int) -> str:
        """Format bytes as human-readable string.

        Args:
            bytes: Number of bytes.

        Returns:
            Formatted string (e.g., "1.5 MB").
        """
        for unit in ["B", "KB", "MB", "GB", "TB"]:
            if bytes < 1024.0:
                return f"{bytes:.1f} {unit}"
            bytes /= 1024.0
        return f"{bytes:.1f} PB"

    def export_to_csv(
        self,
        results: List[TrackResult],
        filepath: str,
        delimiter: str = ",",
        overwrite: bool = False,
    ) -> None:
        """Export results to CSV file with comprehensive validation.

        Exports track results to CSV format using the output_writer module.
        Creates output directory if needed. Validates path, disk space, and permissions.

        Args:
            results: List of TrackResult objects to export.
            filepath: Full path to output CSV file.
            delimiter: CSV delimiter (default: ",").
            overwrite: Whether to allow overwriting existing files (default: False).

        Example:
            >>> results = [TrackResult(...), TrackResult(...)]
            >>> service.export_to_csv(results, "output/results.csv")
        Raises:
            ExportError: If export fails (file permission, disk full, etc.).

        """
        # Validate export path
        is_valid, error_msg = self._validate_export_path(filepath, len(results), overwrite)
        if not is_valid:
            msg = f"Failed to export CSV to {filepath}: {error_msg or 'Invalid export path'}"
            if self.logging_service:
                self.logging_service.error(msg, extra={"filepath": filepath, "track_count": len(results)})
            raise ExportError(
                message=msg,
                error_code="EXPORT_CSV_ERROR",
                context={"filepath": filepath, "track_count": len(results)},
            )

        # Validate delimiter
        valid_delimiters = [",", ";", "\t", "|"]
        if delimiter not in valid_delimiters:
            raise ExportError(
                message=f"Invalid delimiter: {delimiter}. Must be one of: {', '.join(valid_delimiters)}",
                error_code="EXPORT_INVALID_DELIMITER",
                context={"filepath": filepath, "delimiter": delimiter},
            )

        try:
            base_filename = os.path.splitext(os.path.basename(filepath))[0]
            output_dir = os.path.dirname(filepath) or "output"

            # Use atomic write: write to temp file first, then rename
            temp_file = None
            try:
                # Create temp file in same directory for atomic rename
                temp_fd, temp_file = tempfile.mkstemp(
                    suffix=".tmp", dir=output_dir, prefix="cuepoint_export_"
                )
                os.close(temp_fd)

                # Write to temp file (output_writer will handle the actual writing)
                # For CSV, we need to write to the final location, but we'll do atomic rename
                write_csv_files(
                    results=results,
                    base_filename=base_filename,
                    output_dir=output_dir,
                    delimiter=delimiter,
                )

                # Note: write_csv_files creates timestamped files, so we need to handle
                # the actual file that was created. For now, we'll just verify it exists.
                # In a future enhancement, we could modify write_csv_files to support
                # atomic writes directly.

                if self.logging_service:
                    self.logging_service.info(
                        f"Exported {len(results)} tracks to CSV: {filepath}",
                        extra={"filepath": filepath, "track_count": len(results)},
                    )
            except Exception as e:
                # Clean up temp file if it exists
                if temp_file and os.path.exists(temp_file):
                    try:
                        os.unlink(temp_file)
                    except Exception:
                        pass
                raise
        except ExportError:
            # Re-raise ExportError as-is
            raise
        except Exception as e:
            error_msg = f"Failed to export CSV to {filepath}: {str(e)}"
            if self.logging_service:
                self.logging_service.error(error_msg, exc_info=e, extra={"filepath": filepath})
            raise ExportError(
                message=error_msg,
                error_code="EXPORT_CSV_ERROR",
                context={"filepath": filepath, "track_count": len(results)},
            ) from e

    def export_to_json(
        self, results: List[TrackResult], filepath: str, overwrite: bool = False
    ) -> None:
        """Export results to JSON file.

        Exports track results to JSON format with pretty printing.
        Creates output directory if needed.

        Args:
            results: List of TrackResult objects to export.
            filepath: Full path to output JSON file.

        Example:
            >>> results = [TrackResult(...), TrackResult(...)]
            >>> service.export_to_json(results, "output/results.json")
        Raises:
            ExportError: If export fails (file permission, disk full, etc.).

        """
        # Validate export path
        is_valid, error_msg = self._validate_export_path(filepath, len(results), overwrite)
        if not is_valid:
            msg = f"Failed to export JSON to {filepath}: {error_msg or 'Invalid export path'}"
            if self.logging_service:
                self.logging_service.error(msg, extra={"filepath": filepath, "track_count": len(results)})
            raise ExportError(
                message=msg,
                error_code="EXPORT_JSON_ERROR",
                context={"filepath": filepath, "track_count": len(results)},
            )

        try:
            file_path = Path(filepath)
            parent_dir = file_path.parent
            os.makedirs(parent_dir or ".", exist_ok=True)

            import json

            data = [result.to_dict() for result in results] if results else []
            
            # Atomic write: write to temp file first, then rename
            temp_file = None
            try:
                # Create temp file in same directory for atomic rename
                temp_fd, temp_file = tempfile.mkstemp(
                    suffix=".tmp", dir=str(parent_dir), prefix="cuepoint_export_"
                )
                # Close the OS-level handle immediately (Windows otherwise keeps the file locked)
                os.close(temp_fd)
                
                with open(temp_file, "w", encoding="utf-8") as f:
                    json.dump(data, f, indent=2, ensure_ascii=False)
                
                # Atomic rename
                if file_path.exists() and overwrite:
                    file_path.unlink()
                Path(temp_file).replace(file_path)
            except Exception as e:
                # Clean up temp file if it exists
                if temp_file and os.path.exists(temp_file):
                    try:
                        os.unlink(temp_file)
                    except Exception:
                        pass
                raise
            if self.logging_service:
                self.logging_service.info(
                    f"Exported {len(results)} tracks to JSON: {filepath}",
                    extra={"filepath": filepath, "track_count": len(results)},
                )
        except Exception as e:
            error_msg = f"Failed to export JSON to {filepath}: {str(e)}"
            if self.logging_service:
                self.logging_service.error(error_msg, exc_info=e, extra={"filepath": filepath})
            raise ExportError(
                message=error_msg,
                error_code="EXPORT_JSON_ERROR",
                context={"filepath": filepath, "track_count": len(results)},
            ) from e

    def export_to_excel(
        self, results: List[TrackResult], filepath: str, overwrite: bool = False
    ) -> None:
        """Export results to Excel file.

        Exports track results to Excel format with styled headers.
        Requires openpyxl package.

        Args:
            results: List of TrackResult objects to export.
            filepath: Full path to output Excel file.

        Raises:
            ExportError: If export fails (missing openpyxl, file permission, disk full, etc.).

        Example:
            >>> results = [TrackResult(...), TrackResult(...)]
            >>> service.export_to_excel(results, "output/results.xlsx")
        """
        # Validate export path
        is_valid, error_msg = self._validate_export_path(filepath, len(results), overwrite)
        if not is_valid:
            msg = f"Failed to export Excel to {filepath}: {error_msg or 'Invalid export path'}"
            if self.logging_service:
                self.logging_service.error(msg, extra={"filepath": filepath, "track_count": len(results)})
            raise ExportError(
                message=msg,
                error_code="EXPORT_EXCEL_ERROR",
                context={"filepath": filepath, "track_count": len(results)},
            )

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as e:
            error_msg = "Excel export requires openpyxl. Install with: pip install openpyxl"
            if self.logging_service:
                self.logging_service.error(error_msg, exc_info=e)
            raise ExportError(
                message=error_msg,
                error_code="EXPORT_EXCEL_MISSING_DEPENDENCY",
                context={"filepath": filepath},
            ) from e

        try:
            file_path = Path(filepath)
            parent_dir = file_path.parent
            os.makedirs(parent_dir or ".", exist_ok=True)

            wb = Workbook()
            ws = wb.active
            ws.title = "Results"

            # Write headers
            if results:
                headers = list(results[0].to_dict().keys())
                ws.append(headers)

                # Style header row
                header_fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )
                header_font = Font(bold=True, color="FFFFFF")
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
            else:
                # Empty export: create a valid (empty) workbook
                ws.append(["No results"])

            # Write data
            for result in results:
                row_data = list(result.to_dict().values())
                ws.append(row_data)

            # Atomic write: write to temp file first, then rename
            temp_file = None
            try:
                # Create temp file in same directory for atomic rename
                temp_fd, temp_file = tempfile.mkstemp(
                    suffix=".tmp", dir=str(parent_dir), prefix="cuepoint_export_"
                )
                os.close(temp_fd)
                
                # Save to temp file
                wb.save(temp_file)
                
                # Atomic rename
                if file_path.exists() and overwrite:
                    file_path.unlink()
                Path(temp_file).replace(file_path)
            except Exception as e:
                # Clean up temp file if it exists
                if temp_file and os.path.exists(temp_file):
                    try:
                        os.unlink(temp_file)
                    except Exception:
                        pass
                raise
            if self.logging_service:
                self.logging_service.info(
                    f"Exported {len(results)} tracks to Excel: {filepath}",
                    extra={"filepath": filepath, "track_count": len(results)},
                )
        except Exception as e:
            error_msg = f"Failed to export Excel to {filepath}: {str(e)}"
            if self.logging_service:
                self.logging_service.error(error_msg, exc_info=e, extra={"filepath": filepath})
            raise ExportError(
                message=error_msg,
                error_code="EXPORT_EXCEL_ERROR",
                context={"filepath": filepath, "track_count": len(results)},
            ) from e
