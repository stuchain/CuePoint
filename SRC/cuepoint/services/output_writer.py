#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Output Writer Module - CSV file writing functions

This module handles all CSV file writing, separating I/O from business logic.
All functions take TrackResult objects as input and write CSV files.
"""

import csv
import gzip
import json
import os
import time
from datetime import datetime
from typing import Any, Dict, List, Optional, Set

from cuepoint.ui.gui_interface import TrackResult
from cuepoint.utils.utils import with_timestamp

# Try to import openpyxl for Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def write_csv_files(
    results: List[TrackResult],
    base_filename: str,
    output_dir: str = "output",
    delimiter: str = ",",
    include_metadata: bool = True,
) -> Dict[str, str]:
    """
    Write CSV files with custom delimiter.

    Args:
        results: List of TrackResult objects
        base_filename: Base filename (timestamp will be added)
        output_dir: Output directory (default: "output")
        delimiter: CSV delimiter character (default: ",")
        include_metadata: Include metadata columns

    Returns:
        Dictionary mapping file type to file path:
        {
            'main': 'output/filename_20250127_123456.csv',
            'candidates': 'output/filename_candidates_20250127_123456.csv',
            'queries': 'output/filename_queries_20250127_123456.csv',
            'review': 'output/filename_review_20250127_123456.csv' (if needed)
        }

    Raises:
        ValueError: If invalid delimiter provided
    """
    # Validate delimiter
    if delimiter not in [",", ";", "\t", "|"]:
        raise ValueError(f"Invalid delimiter: {delimiter}. Must be one of: , ; \\t |")

    # Ensure output_dir is absolute
    output_dir = os.path.abspath(output_dir)
    os.makedirs(output_dir, exist_ok=True)

    output_files = {}

    # Add timestamp to base filename
    timestamped_filename = with_timestamp(base_filename)

    # Determine file extension based on delimiter
    ext_map = {",": ".csv", ";": ".csv", "\t": ".tsv", "|": ".psv"}
    extension = ext_map.get(delimiter, ".csv")

    # Remove existing extension if present and add correct one
    if timestamped_filename.endswith((".csv", ".tsv", ".psv")):
        timestamped_filename = os.path.splitext(timestamped_filename)[0]
    timestamped_filename = timestamped_filename + extension

    # Write main results CSV
    main_path = write_main_csv(
        results,
        timestamped_filename,
        output_dir,
        delimiter=delimiter,
        include_metadata=include_metadata,
    )
    if main_path:
        output_files["main"] = main_path

    # Write candidates CSV (with same delimiter)
    candidates_path = write_candidates_csv(
        results, timestamped_filename, output_dir, delimiter=delimiter
    )
    if candidates_path:
        output_files["candidates"] = candidates_path

    # Write queries CSV (with same delimiter)
    queries_path = write_queries_csv(results, timestamped_filename, output_dir, delimiter=delimiter)
    if queries_path:
        output_files["queries"] = queries_path

    # Write review CSV (if there are tracks needing review)
    review_indices = _get_review_indices(results)
    if review_indices:
        review_path = write_review_csv(
            results,
            review_indices,
            timestamped_filename,
            output_dir,
            delimiter=delimiter,
            include_metadata=include_metadata,
        )
        if review_path:
            output_files["review"] = review_path

    return output_files


def write_main_csv(
    results: List[TrackResult],
    base_filename: str,
    output_dir: str = "output",
    delimiter: str = ",",
    include_metadata: bool = True,
) -> Optional[str]:
    """
    Write main results CSV file (one row per track) with custom delimiter.

    Args:
        results: List of TrackResult objects
        base_filename: Base filename with timestamp
        output_dir: Output directory
        delimiter: CSV delimiter character (default: ",")
        include_metadata: Include metadata columns

    Returns:
        Path to written file, or None if no results

    Raises:
        OSError: If file cannot be written
        ValueError: If invalid delimiter provided
    """
    if not results:
        return None

    # Validate delimiter
    if delimiter not in [",", ";", "\t", "|"]:
        raise ValueError(f"Invalid delimiter: {delimiter}. Must be one of: , ; \\t |")

    export_start_time = time.time()

    try:
        # Ensure output directory exists
        output_dir = os.path.abspath(output_dir)
        os.makedirs(output_dir, exist_ok=True)

        filepath = os.path.join(output_dir, base_filename)
        filepath = os.path.abspath(filepath)  # Ensure absolute path

        # Define CSV columns
        fieldnames = [
            "playlist_index",
            "original_title",
            "original_artists",
            "beatport_title",
            "beatport_artists",
            "beatport_key",
            "beatport_key_camelot",
            "beatport_year",
            "beatport_bpm",
            "beatport_url",
            "title_sim",
            "artist_sim",
            "match_score",
            "confidence",
            "search_query_index",
            "search_stop_query_index",
            "candidate_index",
        ]

        # Add metadata columns if requested
        if include_metadata:
            fieldnames.extend(
                [
                    "beatport_label",
                    "beatport_genres",
                    "beatport_release",
                    "beatport_release_date",
                    "beatport_track_id",
                ]
            )

        # Write file and ensure it's fully closed before returning
        try:
            with open(filepath, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=delimiter)
                writer.writeheader()
                for result in results:
                    row_dict = result.to_dict()
                    # Filter to only include requested columns
                    filtered_row = {k: row_dict.get(k, "") for k in fieldnames}
                    writer.writerow(filtered_row)
                # Force flush to ensure file is written to disk
                f.flush()
                os.fsync(f.fileno())

            # File should be closed now (context manager)
            # Verify file actually exists and has content
            if not os.path.exists(filepath):
                return None

            file_size = os.path.getsize(filepath)
            if file_size == 0:
                return None

            # Track performance
            export_duration = time.time() - export_start_time

            # Record export metrics (if performance tracking enabled)
            try:
                from performance import performance_collector

                if hasattr(performance_collector, "record_export"):
                    performance_collector.record_export(
                        format="csv",
                        compressed=False,
                        file_size=file_size,
                        duration=export_duration,
                        track_count=len(results),
                    )
            except (ImportError, AttributeError):
                # Performance tracking not available or method doesn't exist
                pass

            return filepath
        except OSError as e:
            raise OSError(f"Failed to write CSV file: {e}")
    except Exception as e:
        raise RuntimeError(f"CSV export failed: {e}") from e


def write_candidates_csv(
    results: List[TrackResult], base_filename: str, output_dir: str = "output", delimiter: str = ","
) -> Optional[str]:
    """
    Write candidates CSV file (all candidates evaluated for all tracks).

    Args:
        results: List of TrackResult objects
        base_filename: Base filename with timestamp
        output_dir: Output directory
        delimiter: CSV delimiter character (default: ",")

    Returns:
        Path to written file, or None if no candidates
    """
    # Collect all candidates from all tracks
    all_candidates = []
    for result in results:
        all_candidates.extend(result.candidates)

    if not all_candidates:
        return None

    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # Determine extension based on delimiter
    ext_map = {",": ".csv", ";": ".csv", "\t": ".tsv", "|": ".psv"}
    extension = ext_map.get(delimiter, ".csv")

    # Remove existing extension and add _candidates with correct extension
    base = os.path.splitext(base_filename)[0]
    filepath = os.path.join(output_dir, f"{base}_candidates{extension}")

    # Get fieldnames from first candidate
    fieldnames = list(all_candidates[0].keys())

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=delimiter)
        writer.writeheader()
        writer.writerows(all_candidates)

    return filepath


def write_queries_csv(
    results: List[TrackResult], base_filename: str, output_dir: str = "output", delimiter: str = ","
) -> Optional[str]:
    """
    Write queries CSV file (all queries executed for all tracks).

    Args:
        results: List of TrackResult objects
        base_filename: Base filename with timestamp
        output_dir: Output directory
        delimiter: CSV delimiter character (default: ",")

    Returns:
        Path to written file, or None if no queries
    """
    # Collect all queries from all tracks
    all_queries = []
    for result in results:
        all_queries.extend(result.queries)

    if not all_queries:
        return None

    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # Determine extension based on delimiter
    ext_map = {",": ".csv", ";": ".csv", "\t": ".tsv", "|": ".psv"}
    extension = ext_map.get(delimiter, ".csv")

    # Remove existing extension and add _queries with correct extension
    base = os.path.splitext(base_filename)[0]
    filepath = os.path.join(output_dir, f"{base}_queries{extension}")

    # Get fieldnames from first query
    fieldnames = list(all_queries[0].keys())

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=delimiter)
        writer.writeheader()
        writer.writerows(all_queries)

    return filepath


def write_review_csv(
    results: List[TrackResult],
    review_indices: Set[int],
    base_filename: str,
    output_dir: str = "output",
    delimiter: str = ",",
    include_metadata: bool = True,
) -> Optional[str]:
    """
    Write review CSV file (tracks needing manual review).

    Args:
        results: List of TrackResult objects
        review_indices: Set of playlist indices that need review
        base_filename: Base filename with timestamp
        output_dir: Output directory
        delimiter: CSV delimiter character (default: ",")
        include_metadata: Include metadata columns

    Returns:
        Path to written file, or None if no review tracks
    """
    review_results = [r for r in results if r.playlist_index in review_indices]
    if not review_results:
        return None

    # Determine extension based on delimiter
    ext_map = {",": ".csv", ";": ".csv", "\t": ".tsv", "|": ".psv"}
    extension = ext_map.get(delimiter, ".csv")

    # Remove existing extension and add _review with correct extension
    base = os.path.splitext(base_filename)[0]
    filepath = os.path.join(output_dir, f"{base}_review{extension}")

    # Use same format as main CSV
    fieldnames = [
        "playlist_index",
        "original_title",
        "original_artists",
        "beatport_title",
        "beatport_artists",
        "beatport_key",
        "beatport_key_camelot",
        "beatport_year",
        "beatport_bpm",
        "beatport_url",
        "title_sim",
        "artist_sim",
        "match_score",
        "confidence",
        "search_query_index",
        "search_stop_query_index",
        "candidate_index",
    ]

    # Add metadata columns if requested
    if include_metadata:
        fieldnames.extend(
            [
                "beatport_label",
                "beatport_genres",
                "beatport_release",
                "beatport_release_date",
                "beatport_track_id",
            ]
        )

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=delimiter)
        writer.writeheader()
        for result in review_results:
            row_dict = result.to_dict()
            # Filter to only include requested columns
            filtered_row = {k: row_dict.get(k, "") for k in fieldnames}
            writer.writerow(filtered_row)

    return filepath


def write_review_candidates_csv(
    results: List[TrackResult],
    review_indices: Set[int],
    base_filename: str,
    output_dir: str = "output",
) -> Optional[str]:
    """Write candidates CSV for review tracks only"""
    review_results = [r for r in results if r.playlist_index in review_indices]
    if not review_results:
        return None

    all_candidates = []
    for result in review_results:
        all_candidates.extend(result.candidates)

    if not all_candidates:
        return None

    base = base_filename.replace(".csv", "") if base_filename.endswith(".csv") else base_filename
    filepath = os.path.join(output_dir, f"{base}_review_candidates.csv")

    fieldnames = list(all_candidates[0].keys())

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_candidates)

    return filepath


def write_review_queries_csv(
    results: List[TrackResult],
    review_indices: Set[int],
    base_filename: str,
    output_dir: str = "output",
) -> Optional[str]:
    """Write queries CSV for review tracks only"""
    review_results = [r for r in results if r.playlist_index in review_indices]
    if not review_results:
        return None

    all_queries = []
    for result in review_results:
        all_queries.extend(result.queries)

    if not all_queries:
        return None

    base = base_filename.replace(".csv", "") if base_filename.endswith(".csv") else base_filename
    filepath = os.path.join(output_dir, f"{base}_review_queries.csv")

    fieldnames = list(all_queries[0].keys())

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_queries)

    return filepath


def _get_review_indices(results: List[TrackResult]) -> Set[int]:
    """
    Determine which tracks need review based on score and match quality.

    Review criteria (from old code):
    - Score < 70
    - Artist similarity < 50 (if artists present)
    - No match found
    """
    review_indices = set()

    for result in results:
        needs_review = False

        # Check score
        if result.match_score is not None and result.match_score < 70:
            needs_review = True

        # Check artist similarity (if artists present)
        if result.artist and result.artist.strip():
            if result.artist_sim is not None and result.artist_sim < 50:
                needs_review = True

        # Check if no match
        if not result.matched or not (result.beatport_url or "").strip():
            needs_review = True

        if needs_review:
            review_indices.add(result.playlist_index)

    return review_indices


def write_json_file(
    results: List[TrackResult],
    file_path: str,
    playlist_name: str = "",
    include_candidates: bool = False,
    include_queries: bool = False,
    include_metadata: bool = True,
    include_processing_info: bool = False,
    compress: bool = False,
    settings: Optional[Dict[str, Any]] = None,
) -> str:
    """
    Write results to JSON file with enhanced options.

    Args:
        results: List of TrackResult objects
        file_path: Full path to output JSON file
        playlist_name: Name of playlist (for metadata)
        include_candidates: Whether to include candidates data
        include_queries: Whether to include queries data
        include_metadata: Include full metadata (genres, labels, etc.)
        include_processing_info: Include processing information
        compress: Compress output using gzip
        settings: Processing settings to include (if include_processing_info is True)

    Returns:
        Path to written file

    Raises:
        OSError: If file cannot be written (permissions, disk full, etc.)
        ValueError: If invalid parameters provided
    """
    # Start performance tracking
    export_start_time = time.time()

    try:
        # Determine if compression is needed based on file extension
        if file_path.endswith(".gz"):
            compress = True
            # Remove .gz extension for base filename
            if file_path.endswith(".json.gz"):
                base_file_path = file_path[:-3]  # Remove .gz
            else:
                base_file_path = file_path[:-3]
        else:
            base_file_path = file_path

        # Ensure directory exists
        output_dir = os.path.dirname(base_file_path) if os.path.dirname(base_file_path) else "."
        os.makedirs(output_dir, exist_ok=True)

        # Build JSON structure
        json_data = {
            "version": "1.0",
            "generated": datetime.now().isoformat(),
            "total_tracks": len(results),
            "matched_tracks": sum(1 for r in results if r.matched),
            "tracks": [],
        }

        # Add processing info if requested
        if include_processing_info:
            json_data["processing_info"] = {
                "timestamp": datetime.now().isoformat(),
                "settings": settings or {},
                "export_format": "json",
                "compressed": compress,
            }

        # Add track data
        for result in results:
            track_data = {
                "playlist_index": result.playlist_index,
                "title": result.title,
                "artist": result.artist or "",
                "matched": result.matched,
            }

            if result.matched and result.beatport_url:
                track_data["match"] = {
                    "beatport_title": result.beatport_title or "",
                    "beatport_artists": result.beatport_artists or "",
                    "beatport_url": result.beatport_url,
                    "match_score": (
                        float(result.match_score) if result.match_score is not None else None
                    ),
                    "confidence": result.confidence or "unknown",
                    "key": result.beatport_key or "",
                    "bpm": (
                        int(result.beatport_bpm)
                        if result.beatport_bpm and result.beatport_bpm.isdigit()
                        else None
                    ),
                    "year": (
                        int(result.beatport_year)
                        if result.beatport_year and result.beatport_year.isdigit()
                        else None
                    ),
                }

                # Include full metadata if requested
                if include_metadata:
                    track_data["match"]["metadata"] = {
                        "label": result.beatport_label or "",
                        "genres": [
                            g.strip()
                            for g in (result.beatport_genres or "").split(",")
                            if g.strip()
                        ],
                        "release": result.beatport_release or "",
                        "release_date": result.beatport_release_date or "",
                        "key_camelot": result.beatport_key_camelot or "",
                        "beatport_track_id": result.beatport_track_id or "",
                    }
                    # Also include scores in metadata section
                    track_data["match"]["scores"] = {
                        "match_score": (
                            float(result.match_score) if result.match_score is not None else None
                        ),
                        "title_sim": (
                            float(result.title_sim) if result.title_sim is not None else None
                        ),
                        "artist_sim": (
                            float(result.artist_sim) if result.artist_sim is not None else None
                        ),
                    }

            # Add candidates if available
            if include_candidates and result.candidates:
                track_data["candidates"] = [
                    {
                        "title": c.get("beatport_title", ""),
                        "artists": c.get("beatport_artists", ""),
                        "url": c.get("beatport_url", ""),
                        "score": c.get("match_score", 0),
                    }
                    for c in result.candidates[:10]  # Top 10 candidates
                ]

            # Add queries if requested
            if include_queries and result.queries:
                track_data["queries"] = result.queries

            json_data["tracks"].append(track_data)

        # Determine final filename
        final_filepath = file_path
        if compress and not final_filepath.endswith(".gz"):
            final_filepath = final_filepath + ".gz"

        # Serialize to JSON
        json_str = json.dumps(json_data, indent=2, ensure_ascii=False)

        # Write to file (with compression if requested)
        try:
            if compress:
                with gzip.open(final_filepath, "wt", encoding="utf-8") as f:
                    f.write(json_str)
            else:
                with open(final_filepath, "w", encoding="utf-8") as f:
                    f.write(json_str)
        except OSError as e:
            raise OSError(f"Failed to write export file: {e}")

        # Track performance
        export_duration = time.time() - export_start_time
        file_size = os.path.getsize(final_filepath)

        # Record export metrics (if performance tracking enabled)
        try:
            from performance import performance_collector

            if hasattr(performance_collector, "record_export"):
                performance_collector.record_export(
                    format="json",
                    compressed=compress,
                    file_size=file_size,
                    duration=export_duration,
                    track_count=len(results),
                )
        except (ImportError, AttributeError):
            # Performance tracking not available or method doesn't exist
            pass

        return final_filepath

    except Exception as e:
        # Log error and re-raise with context
        raise RuntimeError(f"JSON export failed: {e}") from e


def write_excel_file(results: List[TrackResult], file_path: str, playlist_name: str = "") -> str:
    """
    Write results to Excel file (.xlsx).

    Args:
        results: List of TrackResult objects
        file_path: Full path to output Excel file
        playlist_name: Name of playlist (for sheet name)

    Returns:
        Path to written file

    Raises:
        ImportError: If openpyxl is not installed
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl is required for Excel export. " "Install it with: pip install openpyxl"
        )

    os.makedirs(os.path.dirname(file_path) if os.path.dirname(file_path) else ".", exist_ok=True)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = playlist_name[:31] if playlist_name else "Results"  # Excel sheet name limit

    # Define headers
    headers = [
        "Index",
        "Title",
        "Artist",
        "Matched",
        "Beatport Title",
        "Beatport Artist",
        "Score",
        "Confidence",
        "Key",
        "BPM",
        "Year",
        "Label",
        "Genres",
        "Release",
        "URL",
    ]

    # Write headers with styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data rows
    for row_idx, result in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1, value=result.playlist_index)
        ws.cell(row=row_idx, column=2, value=result.title)
        ws.cell(row=row_idx, column=3, value=result.artist or "")
        ws.cell(row=row_idx, column=4, value="Yes" if result.matched else "No")
        ws.cell(row=row_idx, column=5, value=result.beatport_title or "")
        ws.cell(row=row_idx, column=6, value=result.beatport_artists or "")
        ws.cell(
            row=row_idx,
            column=7,
            value=float(result.match_score) if result.match_score is not None else None,
        )
        ws.cell(row=row_idx, column=8, value=result.confidence or "")
        ws.cell(
            row=row_idx, column=9, value=result.beatport_key_camelot or result.beatport_key or ""
        )
        ws.cell(
            row=row_idx,
            column=10,
            value=(
                int(result.beatport_bpm)
                if result.beatport_bpm and result.beatport_bpm.isdigit()
                else None
            ),
        )
        ws.cell(
            row=row_idx,
            column=11,
            value=(
                int(result.beatport_year)
                if result.beatport_year and result.beatport_year.isdigit()
                else None
            ),
        )
        ws.cell(row=row_idx, column=12, value=result.beatport_label or "")
        ws.cell(row=row_idx, column=13, value=result.beatport_genres or "")
        ws.cell(row=row_idx, column=14, value=result.beatport_release or "")
        ws.cell(row=row_idx, column=15, value=result.beatport_url or "")

        # Color code matched/unmatched rows
        if result.matched:
            fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, start=1):
        column_letter = get_column_letter(col_idx)
        max_length = len(header)
        for row_idx in range(2, len(results) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Save workbook
    wb.save(file_path)

    return file_path


def write_performance_report(stats, base_filename: str, output_dir: str = "output") -> str:
    """
    Generate and save performance report to file

    Args:
        stats: PerformanceStats object
        base_filename: Base filename for the report
        output_dir: Output directory (default: "output")

    Returns:
        Path to the generated report file
    """
    from collections import defaultdict

    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # Generate report text
    report_lines = []
    report_lines.append("=" * 80)
    report_lines.append("CuePoint Performance Analysis Report")
    report_lines.append("=" * 80)
    report_lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    report_lines.append("")

    # Overall Statistics
    report_lines.append("Overall Performance:")
    report_lines.append(f"  Total tracks processed: {stats.total_tracks}")
    report_lines.append(f"  Matched tracks: {stats.matched_tracks} ({stats.match_rate():.1f}%)")
    report_lines.append(f"  Unmatched tracks: {stats.unmatched_tracks}")
    report_lines.append(f"  Total processing time: {_format_time_for_report(stats.total_time)}")
    report_lines.append(
        f"  Average time per track: {_format_time_for_report(stats.average_time_per_track())}"
    )
    report_lines.append("")

    # Query Statistics
    report_lines.append("Query Performance:")
    report_lines.append(f"  Total queries executed: {len(stats.query_metrics)}")
    report_lines.append(
        f"  Average time per query: {_format_time_for_report(stats.average_time_per_query())}"
    )
    report_lines.append("")

    # Query Type Breakdown
    by_type = defaultdict(lambda: {"count": 0, "total_time": 0.0, "total_candidates": 0})
    for query in stats.query_metrics:
        qtype = query.query_type
        by_type[qtype]["count"] += 1
        by_type[qtype]["total_time"] += query.execution_time
        by_type[qtype]["total_candidates"] += query.candidates_found

    report_lines.append("Query Performance by Type:")
    for qtype, data in sorted(by_type.items()):
        avg_time = data["total_time"] / data["count"] if data["count"] > 0 else 0.0
        avg_candidates = data["total_candidates"] / data["count"] if data["count"] > 0 else 0.0
        report_lines.append(f"  {qtype.replace('_', ' ').title()}:")
        report_lines.append(f"    Count: {data['count']}")
        report_lines.append(f"    Avg time: {_format_time_for_report(avg_time)}")
        report_lines.append(f"    Avg candidates: {avg_candidates:.1f}")
    report_lines.append("")

    # Cache Statistics
    report_lines.append("Cache Performance:")
    report_lines.append(f"  Cache hits: {stats.cache_stats['hits']}")
    report_lines.append(f"  Cache misses: {stats.cache_stats['misses']}")
    report_lines.append(f"  Hit rate: {stats.cache_hit_rate():.1f}%")
    report_lines.append("")

    # Slowest Tracks
    slowest = sorted(stats.track_metrics, key=lambda t: t.total_time, reverse=True)[:10]
    report_lines.append("Slowest Tracks (Top 10):")
    for track in slowest:
        title_preview = track.track_title[:60]
        time_str = _format_time_for_report(track.total_time)
        report_lines.append(
            f"  {title_preview}: {time_str} ({track.total_queries} queries)"
        )
    report_lines.append("")

    # Bottleneck Analysis
    bottlenecks = _identify_bottlenecks(stats)
    if bottlenecks:
        report_lines.append("Performance Bottlenecks:")
        for bottleneck in bottlenecks:
            report_lines.append(f"  • {bottleneck}")
        report_lines.append("")

    # Write to file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{base_filename}_performance_{timestamp}.txt"
    filepath = os.path.join(output_dir, filename)

    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(report_lines))

    return filepath


def _format_time_for_report(seconds: float) -> str:
    """Format time in seconds to human-readable string for reports"""
    if seconds < 1.0:
        return f"{seconds * 1000:.0f}ms"
    elif seconds < 60.0:
        return f"{seconds:.2f}s"
    else:
        minutes = int(seconds // 60)
        secs = seconds % 60
        return f"{minutes}m {secs:.1f}s"


def _identify_bottlenecks(stats) -> List[str]:
    """Identify performance bottlenecks"""
    bottlenecks = []

    # Check query times
    avg_query_time = stats.average_time_per_query()
    if avg_query_time > 5.0:
        bottlenecks.append(f"Slow queries (avg {avg_query_time:.1f}s per query)")

    # Check cache hit rate
    hit_rate = stats.cache_hit_rate()
    if hit_rate < 30.0 and stats.cache_stats["misses"] > 10:
        bottlenecks.append(f"Low cache hit rate ({hit_rate:.1f}%)")

    # Check track processing time
    avg_track_time = stats.average_time_per_track()
    if avg_track_time > 60.0:
        bottlenecks.append(f"Slow track processing (avg {avg_track_time:.1f}s per track)")

    # Check match rate
    match_rate = stats.match_rate()
    if match_rate < 50.0 and stats.total_tracks > 10:
        bottlenecks.append(f"Low match rate ({match_rate:.1f}%)")

    return bottlenecks
