#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
History View Widget Module - Past searches history viewer

This module contains the HistoryView class for viewing past search results from CSV files.
"""

import csv
import gzip
import json
import os
import time
from datetime import datetime
from typing import Any, Dict, List, Optional

from PySide6.QtCore import Qt, QTimer
from PySide6.QtWidgets import (
    QComboBox,
    QDialog,
    QFileDialog,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMessageBox,
    QPushButton,
    QSpinBox,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from cuepoint.ui.dialogs.export_dialog import ExportDialog
from cuepoint.ui.widgets.candidate_dialog import CandidateDialog

try:
    from performance import performance_collector

    PERFORMANCE_AVAILABLE = True
except ImportError:
    PERFORMANCE_AVAILABLE = False
    performance_collector = None


class HistoryView(QWidget):
    """Widget for viewing past search results from CSV files"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.current_csv_path: Optional[str] = None
        self.csv_rows: List[dict] = []  # Store loaded CSV rows for updates
        self.filtered_rows: List[dict] = []  # Store filtered rows
        self._filter_debounce_timer = QTimer()
        self._filter_debounce_timer.setSingleShot(True)
        self._filter_debounce_timer.timeout.connect(self._apply_filters_debounced)
        self.init_ui()

    def init_ui(self):
        """Initialize UI components"""
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)

        # Create vertical splitter for resizable sections
        splitter = QSplitter(Qt.Vertical)
        splitter.setChildrenCollapsible(False)  # Prevent sections from being collapsed completely

        # Top section: File selection and filters
        top_widget = QWidget()
        top_layout = QVBoxLayout(top_widget)
        top_layout.setSpacing(10)
        top_layout.setContentsMargins(0, 0, 0, 0)

        # File selection section
        file_group = QGroupBox("Select Past Search")
        file_layout = QVBoxLayout()

        # Browse button
        browse_layout = QHBoxLayout()
        self.file_path_label = QLabel("No file selected")
        self.file_path_label.setWordWrap(True)
        browse_layout.addWidget(self.file_path_label, 1)

        self.browse_btn = QPushButton("Browse CSV File...")
        self.browse_btn.clicked.connect(self.on_browse_csv)
        browse_layout.addWidget(self.browse_btn)

        file_layout.addLayout(browse_layout)

        # Recent CSV files list
        recent_label = QLabel("Recent CSV Files:")
        file_layout.addWidget(recent_label)

        self.recent_list = QListWidget()
        self.recent_list.itemDoubleClicked.connect(self.on_recent_file_selected)
        file_layout.addWidget(self.recent_list)

        # Refresh recent files button
        refresh_btn = QPushButton("Refresh List")
        refresh_btn.clicked.connect(self.refresh_recent_files)
        file_layout.addWidget(refresh_btn)

        file_group.setLayout(file_layout)
        top_layout.addWidget(file_group)

        # Results display section (filters only, table goes in bottom section)
        results_group = QGroupBox("Search Results")
        results_layout = QVBoxLayout()

        # Summary info
        self.summary_label = QLabel("No file loaded")
        self.summary_label.setWordWrap(True)
        results_layout.addWidget(self.summary_label)

        # Filter controls
        filter_layout = QHBoxLayout()

        # Search box
        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("Search...")
        self.search_box.textChanged.connect(self._trigger_filter_debounced)
        filter_layout.addWidget(QLabel("Search:"))
        filter_layout.addWidget(self.search_box)

        # Confidence filter
        filter_layout.addWidget(QLabel("Confidence:"))
        self.confidence_filter = QComboBox()
        self.confidence_filter.addItems(["All", "High", "Medium", "Low"])
        self.confidence_filter.currentTextChanged.connect(self._trigger_filter_debounced)
        filter_layout.addWidget(self.confidence_filter)

        filter_layout.addStretch()
        results_layout.addLayout(filter_layout)

        # Advanced Filters Group
        advanced_filters_group = QGroupBox("Advanced Filters")
        advanced_filters_group.setCheckable(True)
        advanced_filters_group.setChecked(False)  # Unchecked by default
        advanced_filters_layout = QVBoxLayout()

        # Container widget for filter controls (to show/hide)
        self.advanced_filters_container = QWidget()
        container_layout = QVBoxLayout(self.advanced_filters_container)
        container_layout.setContentsMargins(0, 0, 0, 0)

        # Year range filter
        year_layout = QHBoxLayout()
        year_layout.addWidget(QLabel("Year Range:"))
        self.year_min = QSpinBox()
        self.year_min.setMinimum(1900)
        self.year_min.setMaximum(2100)
        self.year_min.setValue(1900)
        self.year_min.setSpecialValueText("Any")
        self.year_min.setToolTip("Minimum year (leave at 1900 for no minimum)")
        self.year_min.setSuffix(" -")
        self.year_min.valueChanged.connect(self._trigger_filter_debounced)
        year_layout.addWidget(self.year_min)

        self.year_max = QSpinBox()
        self.year_max.setMinimum(1900)
        self.year_max.setMaximum(2100)
        self.year_max.setValue(2100)
        self.year_max.setSpecialValueText("Any")
        self.year_max.setToolTip("Maximum year (leave at 2100 for no maximum)")
        self.year_max.valueChanged.connect(self._trigger_filter_debounced)
        year_layout.addWidget(self.year_max)
        year_layout.addStretch()
        container_layout.addLayout(year_layout)

        # BPM range filter
        bpm_layout = QHBoxLayout()
        bpm_layout.addWidget(QLabel("BPM Range:"))
        self.bpm_min = QSpinBox()
        self.bpm_min.setMinimum(60)
        self.bpm_min.setMaximum(200)
        self.bpm_min.setValue(60)
        self.bpm_min.setSpecialValueText("Any")
        self.bpm_min.setToolTip("Minimum BPM (leave at 60 for no minimum)")
        self.bpm_min.setSuffix(" -")
        self.bpm_min.valueChanged.connect(self._trigger_filter_debounced)
        bpm_layout.addWidget(self.bpm_min)

        self.bpm_max = QSpinBox()
        self.bpm_max.setMinimum(60)
        self.bpm_max.setMaximum(200)
        self.bpm_max.setValue(200)
        self.bpm_max.setSpecialValueText("Any")
        self.bpm_max.setToolTip("Maximum BPM (leave at 200 for no maximum)")
        self.bpm_max.valueChanged.connect(self._trigger_filter_debounced)
        bpm_layout.addWidget(self.bpm_max)
        bpm_layout.addStretch()
        container_layout.addLayout(bpm_layout)

        # Key filter
        key_layout = QHBoxLayout()
        key_layout.addWidget(QLabel("Musical Key:"))
        self.key_filter = QComboBox()
        keys = (
            ["All"]
            + [f"{k} Major" for k in "C C# D D# E F F# G G# A A# B".split()]
            + [f"{k} Minor" for k in "C C# D D# E F F# G G# A A# B".split()]
        )
        self.key_filter.addItems(keys)
        self.key_filter.setToolTip(
            "Filter by musical key (only shows matched tracks with key data)"
        )
        self.key_filter.setMinimumWidth(150)
        self.key_filter.currentTextChanged.connect(self._trigger_filter_debounced)
        key_layout.addWidget(self.key_filter)
        key_layout.addStretch()
        container_layout.addLayout(key_layout)

        # Clear filters button
        clear_button = QPushButton("Clear All Filters")
        clear_button.setToolTip("Reset all filters to default values")
        clear_button.clicked.connect(self.clear_filters)
        container_layout.addWidget(clear_button)

        # Hide container by default
        self.advanced_filters_container.setVisible(False)

        # Connect checkbox to show/hide container
        advanced_filters_group.toggled.connect(self.advanced_filters_container.setVisible)

        advanced_filters_layout.addWidget(self.advanced_filters_container)
        advanced_filters_group.setLayout(advanced_filters_layout)
        results_layout.addWidget(advanced_filters_group)

        # Result count and filter status
        status_layout = QHBoxLayout()
        self.result_count_label = QLabel("0 results")
        self.result_count_label.setStyleSheet("font-weight: bold;")
        status_layout.addWidget(self.result_count_label)

        self.filter_status_label = QLabel("")
        self.filter_status_label.setStyleSheet("color: gray;")
        status_layout.addStretch()
        status_layout.addWidget(self.filter_status_label)

        # Export button
        self.export_btn = QPushButton("Export...")
        self.export_btn.clicked.connect(self.show_export_dialog)
        self.export_btn.setEnabled(False)  # Disabled until file is loaded
        status_layout.addWidget(self.export_btn)

        results_layout.addLayout(status_layout)

        # Results table (reuse same structure as ResultsView)
        self.table = QTableWidget()
        self.table.setSortingEnabled(True)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        # Enable context menu and double-click for viewing candidates
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._show_context_menu)
        self.table.doubleClicked.connect(self._on_row_double_clicked)

        # Note: filter_layout, summary_label, advanced_filters_group, and status_layout
        # are already added to results_layout above, so don't add them again

        # Don't add table to results_layout - it goes in bottom section
        # results_layout.addWidget(self.table, 1)  # REMOVED - table goes in bottom section

        results_group.setLayout(results_layout)
        top_layout.addWidget(results_group)

        # Add top section to splitter
        top_widget.setMaximumHeight(500)  # Set max height for top section
        splitter.addWidget(top_widget)

        # Bottom section: Results table
        bottom_widget = QWidget()
        bottom_layout = QVBoxLayout(bottom_widget)
        bottom_layout.setContentsMargins(0, 0, 0, 0)

        table_group = QGroupBox("Results Table")
        table_layout = QVBoxLayout()
        table_layout.addWidget(self.table, 1)
        table_group.setLayout(table_layout)
        bottom_layout.addWidget(table_group)

        splitter.addWidget(bottom_widget)

        # Set initial splitter sizes (40% top, 60% bottom)
        splitter.setSizes([400, 600])

        # Add splitter to main layout
        layout.addWidget(splitter, 1)  # Give splitter stretch priority

        # Load recent files on init
        self.refresh_recent_files()

    def on_browse_csv(self):
        """Browse for CSV file to load"""
        # Default to first output directory if it exists
        output_dirs = self._get_output_dirs()
        default_dir = output_dirs[0] if output_dirs else ""

        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select CSV File", default_dir, "CSV Files (*.csv);;All Files (*.*)"
        )

        if file_path:
            self.load_csv_file(file_path)

    def on_recent_file_selected(self, item: QListWidgetItem):
        """Load CSV file from recent list"""
        file_path = item.data(Qt.UserRole)
        if file_path and os.path.exists(file_path):
            self.load_csv_file(file_path)
        else:
            QMessageBox.warning(self, "File Not Found", f"The file no longer exists:\n{file_path}")
            # Refresh the list to remove missing files
            self.refresh_recent_files()

    def load_csv_file(self, file_path: str):
        """Load and display CSV file contents"""
        try:
            if not os.path.exists(file_path):
                QMessageBox.warning(self, "File Not Found", f"File not found: {file_path}")
                return

            # Read CSV file
            with open(file_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                rows = list(reader)

            if not rows:
                QMessageBox.warning(self, "Empty File", "The selected CSV file is empty")
                return

            self.current_csv_path = file_path
            self.file_path_label.setText(f"Loaded: {os.path.basename(file_path)}")

            # Enable export button
            self.export_btn.setEnabled(True)

            # Update summary
            total = len(rows)
            # Check for matched tracks (presence of beatport_url or beatport_title)
            matched = sum(
                1
                for row in rows
                if row.get("beatport_url", "").strip() or row.get("beatport_title", "").strip()
            )
            match_rate = (matched / total * 100) if total > 0 else 0

            # Try to extract timestamp from filename
            timestamp_info = ""
            try:
                # Filename format: playlist_name (dd-mm-yy HH-MM).csv (dashes for Windows
                # compatibility, no colons)
                basename = os.path.basename(file_path)
                if "(" in basename and ")" in basename:
                    # Extract timestamp from parentheses
                    start_idx = basename.rfind("(")
                    end_idx = basename.rfind(")")
                    if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
                        timestamp_str = basename[start_idx + 1 : end_idx]
                        # Try to parse the timestamp (dd-mm-yy HH-MM) - new format with dashes (no
                        # colons)
                        dt = datetime.strptime(timestamp_str, "%d-%m-%y %H-%M")
                        timestamp_info = f"\nSearch Date: {dt.strftime('%d/%m/%y %H:%M')}"
            except BaseException:
                # Fallback: try old format with slashes
                try:
                    # Filename format: playlist_name (dd/mm/yy HH:MM).csv
                    basename = os.path.basename(file_path)
                    if "(" in basename and ")" in basename:
                        # Extract timestamp from parentheses
                        start_idx = basename.rfind("(")
                        end_idx = basename.rfind(")")
                        if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
                            timestamp_str = basename[start_idx + 1 : end_idx]
                            # Try to parse the timestamp (dd/mm/yy HH:MM)
                            dt = datetime.strptime(timestamp_str, "%d/%m/%y %H:%M")
                            timestamp_info = f"\nSearch Date: {dt.strftime('%d/%m/%y %H:%M')}"
                except BaseException:
                    # Fallback to old formats
                    try:
                        # Try old format: playlist_name (YYYY-MM-DD HH:MM:SS).csv
                        basename = os.path.basename(file_path)
                        if "(" in basename and ")" in basename:
                            start_idx = basename.rfind("(")
                            end_idx = basename.rfind(")")
                            if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
                                timestamp_str = basename[start_idx + 1 : end_idx]
                                dt = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")
                                timestamp_info = f"\nSearch Date: {dt.strftime('%d/%m/%y %H:%M')}"
                    except BaseException:
                        # Fallback to very old format: playlist_20250127_123456.csv
                        try:
                            basename = os.path.basename(file_path)
                            parts = basename.replace(".csv", "").split("_")
                            if len(parts) >= 3:
                                date_str = parts[-2]
                                time_str = parts[-1]
                                if len(date_str) == 8 and len(time_str) == 6:
                                    dt = datetime.strptime(
                                        f"{date_str}_{time_str}", "%Y%m%d_%H%M%S"
                                    )
                                    timestamp_info = (
                                        f"\nSearch Date: {dt.strftime('%d/%m/%y %H:%M')}"
                                    )
                        except BaseException:
                            pass

            summary_text = (
                f"File: {os.path.basename(file_path)}\n"
                f"Total tracks: {total}\n"
                f"Matched: {matched} ({match_rate:.1f}%)"
                f"{timestamp_info}"
            )
            self.summary_label.setText(summary_text)

            # Store rows for potential updates
            self.csv_rows = rows
            self.filtered_rows = rows.copy()

            # Populate table
            self.apply_filters()

            # Add to recent files if not already there
            self._add_to_recent_files(file_path)

        except Exception as e:
            QMessageBox.critical(self, "Error Loading File", f"Error loading CSV file:\n{str(e)}")

    def _trigger_filter_debounced(self):
        """Trigger filter with debouncing for performance"""
        self._filter_debounce_timer.stop()
        self._filter_debounce_timer.start(300)

    def _apply_filters_debounced(self):
        """Apply filters after debounce delay"""
        self.apply_filters()

    def _year_in_range(self, row: dict, min_year: Optional[int], max_year: Optional[int]) -> bool:
        """Check if row year is in range"""
        year_str = row.get("beatport_year", "").strip()
        if not year_str:
            return False

        try:
            year = int(year_str)
            if min_year and year < min_year:
                return False
            if max_year and year > max_year:
                return False
            return True
        except (ValueError, TypeError):
            return False

    def _bpm_in_range(self, row: dict, min_bpm: Optional[int], max_bpm: Optional[int]) -> bool:
        """Check if row BPM is in range"""
        bpm_str = row.get("beatport_bpm", "").strip()
        if not bpm_str:
            return False

        try:
            bpm = float(bpm_str)
            if min_bpm and bpm < min_bpm:
                return False
            if max_bpm and bpm > max_bpm:
                return False
            return True
        except (ValueError, TypeError):
            return False

    def _update_filter_status(self):
        """Update filter status label to show active filters"""
        active_filters = []

        if self.year_min.value() > 1900 or self.year_max.value() < 2100:
            min_val = self.year_min.value() if self.year_min.value() > 1900 else None
            max_val = self.year_max.value() if self.year_max.value() < 2100 else None
            if min_val and max_val:
                active_filters.append(f"Year: {min_val}-{max_val}")
            elif min_val:
                active_filters.append(f"Year: ≥{min_val}")
            elif max_val:
                active_filters.append(f"Year: ≤{max_val}")

        if self.bpm_min.value() > 60 or self.bpm_max.value() < 200:
            min_val = self.bpm_min.value() if self.bpm_min.value() > 60 else None
            max_val = self.bpm_max.value() if self.bpm_max.value() < 200 else None
            if min_val and max_val:
                active_filters.append(f"BPM: {min_val}-{max_val}")
            elif min_val:
                active_filters.append(f"BPM: ≥{min_val}")
            elif max_val:
                active_filters.append(f"BPM: ≤{max_val}")

        if self.key_filter.currentText() != "All":
            active_filters.append(f"Key: {self.key_filter.currentText()}")

        if self.search_box.text():
            active_filters.append("Search")

        if self.confidence_filter.currentText() != "All":
            active_filters.append(f"Confidence: {self.confidence_filter.currentText()}")

        if active_filters:
            self.filter_status_label.setText(f"Active filters: {', '.join(active_filters)}")
        else:
            self.filter_status_label.setText("No filters active")

    def _filter_rows(self) -> List[dict]:
        """Apply filters to CSV rows"""
        filter_start_time = time.time()

        filtered = self.csv_rows.copy()
        initial_count = len(filtered)

        # Search filter
        search_text = self.search_box.text().lower().strip()
        if search_text:
            filtered = [
                r
                for r in filtered
                if search_text in (r.get("original_title", "") or "").lower()
                or search_text in (r.get("original_artists", "") or "").lower()
                or search_text in (r.get("beatport_title", "") or "").lower()
                or search_text in (r.get("beatport_artists", "") or "").lower()
            ]

        # Confidence filter
        confidence = self.confidence_filter.currentText()
        if confidence != "All":
            filtered = [
                r for r in filtered if (r.get("confidence", "") or "").lower() == confidence.lower()
            ]

        # Year range filter
        year_min_val = self.year_min.value() if self.year_min.value() > 1900 else None
        year_max_val = self.year_max.value() if self.year_max.value() < 2100 else None

        if year_min_val or year_max_val:
            filtered = [r for r in filtered if self._year_in_range(r, year_min_val, year_max_val)]

        # BPM range filter
        bpm_min_val = self.bpm_min.value() if self.bpm_min.value() > 60 else None
        bpm_max_val = self.bpm_max.value() if self.bpm_max.value() < 200 else None

        if bpm_min_val or bpm_max_val:
            filtered = [r for r in filtered if self._bpm_in_range(r, bpm_min_val, bpm_max_val)]

        # Key filter
        key_filter_val = self.key_filter.currentText()
        if key_filter_val != "All":
            filtered = [
                r for r in filtered if (r.get("beatport_key", "") or "").strip() == key_filter_val
            ]

        self.filtered_rows = filtered

        # Track performance
        filter_duration = time.time() - filter_start_time
        if PERFORMANCE_AVAILABLE and performance_collector:
            performance_collector.record_filter_operation(
                duration=filter_duration,
                initial_count=initial_count,
                filtered_count=len(filtered),
                filters_applied={
                    "search": bool(search_text),
                    "confidence": confidence if confidence != "All" else None,
                    "year_range": (year_min_val, year_max_val),
                    "bpm_range": (bpm_min_val, bpm_max_val),
                    "key": key_filter_val if key_filter_val != "All" else None,
                },
            )

        return filtered

    def apply_filters(self):
        """Apply filters and update table"""
        filtered = self._filter_rows()
        self._populate_table(filtered)
        self._update_filter_status()

        # Update result count
        count_text = f"{len(self.filtered_rows)} of {len(self.csv_rows)} results"
        if len(self.filtered_rows) < len(self.csv_rows):
            count_text += " (filtered)"
        self.result_count_label.setText(count_text)

        # Show message if no results
        if len(self.filtered_rows) == 0 and len(self.csv_rows) > 0:
            self.result_count_label.setStyleSheet("font-weight: bold; color: red;")
            self.result_count_label.setText("No results match the current filters")
        else:
            self.result_count_label.setStyleSheet("font-weight: bold;")

    def clear_filters(self):
        """Clear all filters to default values"""
        self.year_min.setValue(1900)
        self.year_max.setValue(2100)
        self.bpm_min.setValue(60)
        self.bpm_max.setValue(200)
        self.key_filter.setCurrentText("All")
        self.search_box.clear()
        self.confidence_filter.setCurrentText("All")
        self.apply_filters()

    def _populate_table(self, rows: List[dict]):
        """Populate table with CSV data"""
        # Store current sort state
        sort_column = self.table.horizontalHeader().sortIndicatorSection()
        sort_order = self.table.horizontalHeader().sortIndicatorOrder()

        if not rows:
            self.table.setRowCount(0)
            return

        # Get column names from first row
        columns = list(rows[0].keys())

        # Set up table
        self.table.setSortingEnabled(False)
        self.table.setColumnCount(len(columns))
        self.table.setHorizontalHeaderLabels(columns)
        self.table.setRowCount(len(rows))

        # Populate rows
        # Find index column and calculate padding
        index_col = -1
        for col_idx, col_name in enumerate(columns):
            if "index" in col_name.lower() or "playlist_index" in col_name.lower():
                index_col = col_idx
                break

        # Calculate padding for index column if found
        padding = 1
        if index_col >= 0:
            max_index = 0
            for row_data in rows:
                try:
                    idx_val = int(row_data.get(columns[index_col], 0) or 0)
                    max_index = max(max_index, idx_val)
                except (ValueError, TypeError):
                    pass
            padding = len(str(max_index)) if max_index > 0 else 1

        for row_idx, row_data in enumerate(rows):
            for col_idx, col_name in enumerate(columns):
                value = row_data.get(col_name, "")
                # Pad index column with zeros for proper sorting
                if col_idx == index_col:
                    try:
                        numeric_value = int(value) if value else 0
                        index_str = str(numeric_value).zfill(padding)
                        item = QTableWidgetItem(index_str)
                        item.setData(Qt.EditRole, numeric_value)
                        item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    except (ValueError, TypeError):
                        item = QTableWidgetItem(str(value))
                else:
                    item = QTableWidgetItem(str(value))
                self.table.setItem(row_idx, col_idx, item)

        # Re-enable sorting
        self.table.setSortingEnabled(True)

        # Find the index column
        index_col = -1
        for col in range(self.table.columnCount()):
            header = self.table.horizontalHeaderItem(col)
            if header:
                header_text = header.text().lower()
                if "index" in header_text or "playlist_index" in header_text:
                    index_col = col
                    break

        # Always default to Index column in ascending order
        # Only restore user's sort if they sorted by a different column
        if index_col >= 0:
            if sort_column >= 0 and sort_column != index_col and sort_order is not None:
                # User sorted by a different column, restore that
                self.table.sortItems(sort_column, sort_order)
            else:
                # Default: sort by Index column in ascending order
                # Clear any previous sort indicator first
                self.table.horizontalHeader().setSortIndicator(-1, Qt.AscendingOrder)
                self.table.sortItems(index_col, Qt.AscendingOrder)
                self.table.horizontalHeader().setSortIndicator(index_col, Qt.AscendingOrder)

        # Resize columns to content
        self.table.resizeColumnsToContents()

        # Set minimum column widths
        for col in range(self.table.columnCount()):
            current_width = self.table.columnWidth(col)
            self.table.setColumnWidth(col, max(current_width, 80))

    def _on_candidate_selected(
        self,
        row: int,
        playlist_index: int,
        original_title: str,
        original_artists: str,
        candidate: Dict[str, Any],
    ):
        """Handle candidate selection - update the CSV row and table display"""
        if row < 0 or row >= len(self.filtered_rows):
            return

        # Get the filtered row
        filtered_row = self.filtered_rows[row]

        # Find the corresponding row in csv_rows by matching key fields
        csv_row = None
        for csv_r in self.csv_rows:
            if (
                csv_r.get("playlist_index") == filtered_row.get("playlist_index")
                and csv_r.get("original_title", "").strip()
                == filtered_row.get("original_title", "").strip()
                and csv_r.get("original_artists", "").strip()
                == filtered_row.get("original_artists", "").strip()
            ):
                csv_row = csv_r
                break

        if csv_row is None:
            return

        # Update match information
        csv_row["beatport_title"] = candidate.get(
            "candidate_title", candidate.get("beatport_title", "")
        )
        csv_row["beatport_artists"] = candidate.get(
            "candidate_artists", candidate.get("beatport_artists", "")
        )
        csv_row["beatport_url"] = candidate.get("beatport_url", candidate.get("candidate_url", ""))
        csv_row["beatport_key"] = candidate.get("beatport_key", candidate.get("candidate_key", ""))
        csv_row["beatport_key_camelot"] = candidate.get(
            "beatport_key_camelot", candidate.get("candidate_key_camelot", "")
        )
        csv_row["beatport_year"] = candidate.get(
            "beatport_year", candidate.get("candidate_year", "")
        )
        csv_row["beatport_bpm"] = candidate.get("beatport_bpm", candidate.get("candidate_bpm", ""))
        csv_row["beatport_label"] = candidate.get(
            "beatport_label", candidate.get("candidate_label", "")
        )
        csv_row["title_sim"] = str(candidate.get("title_sim", ""))
        csv_row["artist_sim"] = str(candidate.get("artist_sim", ""))
        csv_row["match_score"] = str(candidate.get("match_score", candidate.get("final_score", "")))

        # Update confidence based on score
        try:
            score = float(csv_row.get("match_score", 0))
            if score >= 90:
                csv_row["confidence"] = "high"
            elif score >= 75:
                csv_row["confidence"] = "medium"
            else:
                csv_row["confidence"] = "low"
        except (ValueError, TypeError):
            csv_row["confidence"] = ""

        # Update candidate_index if available
        csv_row["candidate_index"] = candidate.get("candidate_index", "")

        # Re-apply filters to update the display
        self.apply_filters()

        # Show confirmation
        QMessageBox.information(
            self, "Candidate Updated", f"Updated match for:\n{original_title} - {original_artists}"
        )

    def _update_table_row(self, row: int, csv_row: dict):
        """Update a specific row in the table with new CSV data"""
        # Find columns by header name
        for col in range(self.table.columnCount()):
            header = self.table.horizontalHeaderItem(col)
            if header:
                col_name = header.text()
                # Map column names to CSV row keys
                value = ""
                if col_name == "Beatport Title":
                    value = csv_row.get("beatport_title", "")
                elif col_name == "Beatport Artist" or col_name == "Beatport Artists":
                    value = csv_row.get("beatport_artists", "")
                elif col_name == "Score":
                    value = csv_row.get("match_score", "")
                elif col_name == "Confidence":
                    value = csv_row.get("confidence", "").capitalize()
                elif col_name == "Key":
                    value = csv_row.get("beatport_key", "")
                elif col_name == "BPM":
                    value = csv_row.get("beatport_bpm", "")
                elif col_name == "Year":
                    value = csv_row.get("beatport_year", "")
                elif col_name == "Title Sim":
                    value = csv_row.get("title_sim", "")
                elif col_name == "Artist Sim":
                    value = csv_row.get("artist_sim", "")
                elif col_name == "Matched":
                    # Update matched status
                    matched = bool(
                        csv_row.get("beatport_url", "").strip()
                        or csv_row.get("beatport_title", "").strip()
                    )
                    matched_item = QTableWidgetItem("✓" if matched else "✗")
                    matched_item.setTextAlignment(Qt.AlignCenter)
                    if matched:
                        matched_item.setForeground(Qt.darkGreen)
                    else:
                        matched_item.setForeground(Qt.darkRed)
                    self.table.setItem(row, col, matched_item)
                    continue
                else:
                    # Try direct column name match (lowercase, replace spaces with underscores)
                    key = col_name.lower().replace(" ", "_")
                    value = csv_row.get(key, "")

                if value or col_name in [
                    "Beatport Title",
                    "Beatport Artist",
                    "Beatport Artists",
                    "Score",
                    "Confidence",
                    "Key",
                    "BPM",
                    "Year",
                ]:
                    item = QTableWidgetItem(str(value))
                    self.table.setItem(row, col, item)

        # Update summary
        self._update_summary()

    def _update_summary(self):
        """Update summary statistics after candidate update"""
        if not self.csv_rows:
            return

        total = len(self.csv_rows)
        matched = sum(
            1
            for row in self.csv_rows
            if row.get("beatport_url", "").strip() or row.get("beatport_title", "").strip()
        )
        match_rate = (matched / total * 100) if total > 0 else 0

        # Try to extract timestamp from filename
        timestamp_info = ""
        if self.current_csv_path:
            try:
                basename = os.path.basename(self.current_csv_path)
                if "(" in basename and ")" in basename:
                    start_idx = basename.rfind("(")
                    end_idx = basename.rfind(")")
                    if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
                        timestamp_str = basename[start_idx + 1 : end_idx]
                        dt = datetime.strptime(timestamp_str, "%d-%m-%y %H-%M")
                        timestamp_info = f"\nSearch Date: {dt.strftime('%d/%m/%y %H:%M')}"
            except BaseException:
                pass

        file_basename = (
            os.path.basename(self.current_csv_path)
            if self.current_csv_path
            else 'No file'
        )
        summary_text = (
            f"File: {file_basename}\n"
            f"Total tracks: {total}\n"
            f"Matched: {matched} ({match_rate:.1f}%)"
            f"{timestamp_info}"
        )
        self.summary_label.setText(summary_text)

    def _get_candidates_csv_path(self, main_csv_path: str) -> Optional[str]:
        """Get the path to the candidates CSV file for a given main CSV file"""
        # Replace main CSV filename with candidates CSV filename
        base_path = main_csv_path.replace(".csv", "")
        candidates_path = f"{base_path}_candidates.csv"
        if os.path.exists(candidates_path):
            return candidates_path
        return None

    def _load_candidates_for_track(
        self, playlist_index: int, original_title: str, original_artists: str
    ) -> List[Dict[str, Any]]:
        """Load candidates for a specific track from the candidates CSV"""
        if not self.current_csv_path:
            return []

        candidates_path = self._get_candidates_csv_path(self.current_csv_path)
        if not candidates_path:
            return []

        try:
            candidates = []
            with open(candidates_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    # Match by playlist_index and original title/artists
                    if (
                        row.get("playlist_index") == str(playlist_index)
                        and row.get("original_title", "").strip() == original_title.strip()
                        and row.get("original_artists", "").strip() == original_artists.strip()
                    ):
                        # Convert CSV row to candidate dict format expected by CandidateDialog
                        candidate = {
                            "candidate_title": row.get("candidate_title", ""),
                            "candidate_artists": row.get("candidate_artists", ""),
                            "beatport_url": row.get("candidate_url", ""),
                            "match_score": (
                                float(row.get("final_score", 0)) if row.get("final_score") else 0.0
                            ),
                            "title_sim": (
                                float(row.get("title_sim", 0)) if row.get("title_sim") else 0.0
                            ),
                            "artist_sim": (
                                float(row.get("artist_sim", 0)) if row.get("artist_sim") else 0.0
                            ),
                            "beatport_key": row.get("candidate_key", ""),
                            "beatport_key_camelot": row.get("candidate_key_camelot", ""),
                            "beatport_bpm": row.get("candidate_bpm", ""),
                            "beatport_year": row.get("candidate_year", ""),
                            "beatport_label": row.get("candidate_label", ""),
                        }
                        candidates.append(candidate)

            # Sort by final_score (descending) to rank them
            candidates.sort(key=lambda x: x.get("match_score", 0), reverse=True)
            return candidates
        except Exception:
            return []

    def _show_context_menu(self, position):
        """Show context menu for table row"""
        from PySide6.QtWidgets import QMenu

        item = self.table.itemAt(position)
        if not item:
            return

        row = item.row()
        menu = QMenu(self)

        # Get track info to check if candidates exist
        playlist_index_item = self.table.item(row, 0)
        if playlist_index_item:
            try:
                playlist_index = int(playlist_index_item.text())
                title_item = self.table.item(row, 1)
                artist_item = self.table.item(row, 2) if self.table.columnCount() > 2 else None

                if title_item:
                    original_title = title_item.text()
                    original_artists = artist_item.text() if artist_item else ""

                    # Check if candidates CSV exists and has candidates for this track
                    candidates = self._load_candidates_for_track(
                        playlist_index, original_title, original_artists
                    )

                    if candidates:
                        view_action = menu.addAction("View Candidates...")
                        view_action.triggered.connect(
                            lambda: self._on_row_double_clicked(self.table.indexFromItem(item))
                        )
                    else:
                        no_candidates_action = menu.addAction("No candidates available")
                        no_candidates_action.setEnabled(False)
            except (ValueError, IndexError):
                pass

        if menu.actions():
            menu.exec(self.table.viewport().mapToGlobal(position))

    def _on_row_double_clicked(self, index):
        """Handle double-click on table row to view candidates"""
        row = index.row()
        if row < 0 or row >= self.table.rowCount():
            return

        # Get track information from the table
        playlist_index_item = self.table.item(row, 0)  # Assuming first column is playlist_index
        title_item = self.table.item(row, 1)  # Assuming second column is title
        artist_item = self.table.item(row, 2)  # Assuming third column is artist

        if not playlist_index_item or not title_item:
            return

        playlist_index = int(playlist_index_item.text())
        original_title = title_item.text()
        original_artists = artist_item.text() if artist_item else ""

        # Load candidates for this track
        candidates = self._load_candidates_for_track(
            playlist_index, original_title, original_artists
        )

        if not candidates:
            QMessageBox.information(
                self,
                "No Candidates",
                "No candidates available for this track.\n\n"
                "The candidates CSV file may not exist or this track "
                "had no candidates during the search.",
            )
            return

        # Get current match info if track is matched
        current_match = None
        # Check if there's a beatport_url or beatport_title in the row
        # We need to find which column has the beatport data
        beatport_title_item = None
        for col in range(self.table.columnCount()):
            header = self.table.horizontalHeaderItem(col)
            if header:
                header_text = header.text().lower()
                if "beatport" in header_text and "title" in header_text:
                    beatport_title_item = self.table.item(row, col)
                elif "beatport" in header_text and "url" in header_text:
                    _ = self.table.item(row, col)  # beatport_url_item unused

        if beatport_title_item and beatport_title_item.text().strip():
            # Find the matched candidate
            beatport_title = beatport_title_item.text().strip()
            for candidate in candidates:
                if candidate.get("candidate_title", "").strip() == beatport_title:
                    current_match = candidate
                    break

        # Show candidate dialog
        dialog = CandidateDialog(
            track_title=original_title,
            track_artist=original_artists,
            candidates=candidates,
            current_match=current_match,
            parent=self,
        )

        # Connect candidate selection to update handler
        dialog.candidate_selected.connect(
            lambda candidate: self._on_candidate_selected(
                row, playlist_index, original_title, original_artists, candidate
            )
        )

        dialog.exec()

    def _get_output_dirs(self):
        """Get list of output directory paths (only SRC/output folder)"""
        # Determine SRC directory
        current_file = os.path.abspath(__file__)  # SRC/gui/history_view.py
        src_dir = os.path.dirname(os.path.dirname(current_file))  # SRC/

        output_dirs = []
        # Only use SRC output folder
        src_output = os.path.join(src_dir, "output")
        if os.path.exists(src_output):
            output_dirs.append(src_output)

        return output_dirs

    def refresh_recent_files(self):
        """Refresh list of recent CSV files from output directories"""
        self.recent_list.clear()

        output_dirs = self._get_output_dirs()
        if not output_dirs:
            return

        # Find all CSV files from all output directories
        csv_files = []
        for output_dir in output_dirs:
            if not os.path.exists(output_dir):
                continue

            # Use glob to find files, which may bypass some caching issues
            import glob

            pattern = os.path.join(output_dir, "*.csv")
            all_csv_files = glob.glob(pattern)

            # Also try os.listdir as fallback
            try:
                dir_files = os.listdir(output_dir)
            except Exception:
                dir_files = []

            # Combine both methods and use set to deduplicate
            found_files = set()
            for file_path in all_csv_files:
                filename = os.path.basename(file_path)
                if (
                    not filename.endswith("_candidates.csv")
                    and not filename.endswith("_queries.csv")
                    and not filename.endswith("_review.csv")
                ):
                    found_files.add(file_path)

            # Also check files from os.listdir
            for filename in dir_files:
                if filename.endswith(".csv"):
                    if (
                        not filename.endswith("_candidates.csv")
                        and not filename.endswith("_queries.csv")
                        and not filename.endswith("_review.csv")
                    ):
                        file_path = os.path.join(output_dir, filename)
                        found_files.add(file_path)

            # Add to csv_files list with modification time
            for file_path in found_files:
                if os.path.exists(file_path):
                    try:
                        mtime = os.path.getmtime(file_path)
                        csv_files.append((file_path, mtime))
                    except Exception:
                        pass

        # Sort by modification time (newest first)
        csv_files.sort(key=lambda x: x[1], reverse=True)

        # Add to list (show last 20)
        for file_path, mtime in csv_files[:20]:
            item = QListWidgetItem(os.path.basename(file_path))
            item.setData(Qt.UserRole, file_path)
            # Add timestamp tooltip with folder info
            dt = datetime.fromtimestamp(mtime)
            folder_name = os.path.basename(os.path.dirname(file_path))
            item.setToolTip(
                f"{file_path}\nFolder: {folder_name}\nModified: {dt.strftime('%Y-%m-%d %H:%M:%S')}"
            )
            self.recent_list.addItem(item)

    def _add_to_recent_files(self, file_path: str):
        """Add file to recent files (for quick access)"""
        # This could use QSettings to persist across sessions
        # For now, just refresh the list
        self.refresh_recent_files()

        # Highlight the selected file in the list
        for i in range(self.recent_list.count()):
            item = self.recent_list.item(i)
            if item.data(Qt.UserRole) == file_path:
                self.recent_list.setCurrentItem(item)
                self.recent_list.scrollToItem(item)
                break

    def show_export_dialog(self):
        """Show export dialog and handle export for CSV row data"""
        if not self.csv_rows:
            QMessageBox.warning(
                self, "No Results", "No results to export. Please load a CSV file first."
            )
            return

        # Get rows to export (filtered or all)
        dialog = ExportDialog(self)
        if dialog.exec() != QDialog.Accepted:
            return

        options = dialog.get_export_options()
        file_path = options.get("file_path")

        if not file_path:
            QMessageBox.warning(self, "No File Selected", "Please select a file location")
            return

        # Get rows to export
        rows_to_export = (
            self.filtered_rows if options.get("export_filtered", False) else self.csv_rows
        )

        if not rows_to_export:
            QMessageBox.warning(
                self, "No Results", "No results to export (filter may have excluded all results)"
            )
            return

        try:
            format_type = options.get("format", "csv")

            if format_type == "json":
                self._export_to_json(rows_to_export, file_path, options)
            elif format_type == "excel":
                self._export_to_excel(rows_to_export, file_path, options)
            else:  # CSV
                self._export_to_csv(rows_to_export, file_path, options)

        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Error exporting file:\n{str(e)}")

    def _export_to_csv(self, rows: List[dict], file_path: str, options: Dict[str, Any]):
        """Export CSV row dictionaries to CSV file"""
        delimiter = options.get("delimiter", ",")
        include_metadata = options.get("include_metadata", True)

        # Validate delimiter
        if delimiter not in [",", ";", "\t", "|"]:
            raise ValueError(f"Invalid delimiter: {delimiter}")

        # Determine file extension
        ext_map = {",": ".csv", ";": ".csv", "\t": ".tsv", "|": ".psv"}
        extension = ext_map.get(delimiter, ".csv")

        # Ensure correct extension
        if not file_path.endswith(extension):
            file_path = os.path.splitext(file_path)[0] + extension

        # Get all column names from rows
        if not rows:
            raise ValueError("No rows to export")

        columns = list(rows[0].keys())

        # Filter columns if metadata not included
        if not include_metadata:
            metadata_cols = {
                "label",
                "genres",
                "release",
                "release_date",
                "beatport_label",
                "beatport_genres",
                "beatport_release",
                "beatport_release_date",
            }
            columns = [col for col in columns if col.lower() not in metadata_cols]

        # Write CSV file
        os.makedirs(os.path.dirname(file_path) or ".", exist_ok=True)

        with open(file_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(
                f, fieldnames=columns, delimiter=delimiter, extrasaction="ignore"
            )
            writer.writeheader()
            writer.writerows(rows)

        QMessageBox.information(self, "Export Complete", f"CSV file exported to:\n{file_path}")

    def _export_to_json(self, rows: List[dict], file_path: str, options: Dict[str, Any]):
        """Export CSV row dictionaries to JSON file"""
        include_metadata = options.get("include_metadata", True)
        include_processing_info = options.get("include_processing_info", False)
        compress = options.get("compress", False)

        # Build JSON structure
        json_data = {
            "version": "1.0",
            "generated": datetime.now().isoformat(),
            "source_file": self.current_csv_path or "unknown",
            "total_tracks": len(rows),
            "matched_tracks": sum(
                1
                for r in rows
                if r.get("beatport_url", "").strip() or r.get("beatport_title", "").strip()
            ),
            "tracks": [],
        }

        # Add processing info if requested
        if include_processing_info:
            json_data["processing_info"] = {
                "timestamp": datetime.now().isoformat(),
                "export_format": "json",
                "compressed": compress,
                "source_csv": self.current_csv_path,
            }

        # Convert rows to track data
        for row in rows:
            track_data = dict(row)  # Copy all fields

            # Filter metadata if not included
            if not include_metadata:
                metadata_keys = [
                    "label",
                    "genres",
                    "release",
                    "release_date",
                    "beatport_label",
                    "beatport_genres",
                    "beatport_release",
                    "beatport_release_date",
                ]
                track_data = {k: v for k, v in track_data.items() if k.lower() not in metadata_keys}

            json_data["tracks"].append(track_data)

        # Determine filename
        if compress and not file_path.endswith(".gz"):
            file_path += ".gz"
        elif not compress and file_path.endswith(".gz"):
            file_path = file_path[:-3]

        # Ensure .json extension
        if not file_path.endswith(".json") and not file_path.endswith(".json.gz"):
            file_path = os.path.splitext(file_path)[0] + ".json"
            if compress:
                file_path += ".gz"

        # Write JSON file
        os.makedirs(os.path.dirname(file_path) or ".", exist_ok=True)
        json_str = json.dumps(json_data, indent=2, ensure_ascii=False)

        if compress:
            with gzip.open(file_path, "wt", encoding="utf-8") as f:
                f.write(json_str)
        else:
            with open(file_path, "w", encoding="utf-8") as f:
                f.write(json_str)

        QMessageBox.information(self, "Export Complete", f"JSON file exported to:\n{file_path}")

    def _export_to_excel(self, rows: List[dict], file_path: str, options: Dict[str, Any]):
        """Export CSV row dictionaries to Excel file"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            QMessageBox.warning(
                self,
                "Missing Dependency",
                "Excel export requires openpyxl. Please install it:\npip install openpyxl",
            )
            return

        include_metadata = options.get("include_metadata", True)

        # Get all column names
        if not rows:
            raise ValueError("No rows to export")

        columns = list(rows[0].keys())

        # Filter columns if metadata not included
        if not include_metadata:
            metadata_cols = {
                "label",
                "genres",
                "release",
                "release_date",
                "beatport_label",
                "beatport_genres",
                "beatport_release",
                "beatport_release_date",
            }
            columns = [col for col in columns if col.lower() not in metadata_cols]

        # Ensure .xlsx extension
        if not file_path.endswith(".xlsx"):
            file_path = os.path.splitext(file_path)[0] + ".xlsx"

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Search Results"

        # Write headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = row_data.get(col_name, "")
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for col_idx, col_name in enumerate(columns, 1):
            max_length = len(str(col_name))
            for row_idx in range(2, len(rows) + 2):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
                max_length + 2, 50
            )

        # Save file
        os.makedirs(os.path.dirname(file_path) or ".", exist_ok=True)
        wb.save(file_path)

        QMessageBox.information(self, "Export Complete", f"Excel file exported to:\n{file_path}")
