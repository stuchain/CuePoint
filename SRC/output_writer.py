#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Output Writer Module - CSV file writing functions

This module handles all CSV file writing, separating I/O from business logic.
All functions take TrackResult objects as input and write CSV files.
"""

import csv
import json
import os
from typing import List, Dict, Set, Optional, Any
from datetime import datetime
from gui_interface import TrackResult
from utils import with_timestamp

# Try to import openpyxl for Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def write_csv_files(
    results: List[TrackResult],
    base_filename: str,
    output_dir: str = "output"
) -> Dict[str, str]:
    """
    Write all CSV files from TrackResult objects.
    
    Args:
        results: List of TrackResult objects
        base_filename: Base filename (timestamp will be added)
        output_dir: Output directory (default: "output")
    
    Returns:
        Dictionary mapping file type to file path:
        {
            'main': 'output/filename_20250127_123456.csv',
            'candidates': 'output/filename_candidates_20250127_123456.csv',
            'queries': 'output/filename_queries_20250127_123456.csv',
            'review': 'output/filename_review_20250127_123456.csv' (if needed)
        }
    """
    os.makedirs(output_dir, exist_ok=True)
    output_files = {}
    
    # Add timestamp to base filename
    timestamped_filename = with_timestamp(base_filename)
    
    # Write main results CSV
    main_path = write_main_csv(results, timestamped_filename, output_dir)
    if main_path:
        output_files['main'] = main_path
    
    # Write candidates CSV
    candidates_path = write_candidates_csv(results, timestamped_filename, output_dir)
    if candidates_path:
        output_files['candidates'] = candidates_path
    
    # Write queries CSV
    queries_path = write_queries_csv(results, timestamped_filename, output_dir)
    if queries_path:
        output_files['queries'] = queries_path
    
    # Write review CSV (if there are tracks needing review)
    review_indices = _get_review_indices(results)
    if review_indices:
        review_path = write_review_csv(results, review_indices, timestamped_filename, output_dir)
        if review_path:
            output_files['review'] = review_path
    
    return output_files


def write_main_csv(
    results: List[TrackResult],
    base_filename: str,
    output_dir: str = "output"
) -> Optional[str]:
    """
    Write main results CSV file (one row per track).
    
    Args:
        results: List of TrackResult objects
        base_filename: Base filename with timestamp
        output_dir: Output directory
    
    Returns:
        Path to written file, or None if no results
    """
    if not results:
        return None
    
    filepath = os.path.join(output_dir, base_filename)
    
    # Define CSV columns (must match old format exactly)
    fieldnames = [
        "playlist_index", "original_title", "original_artists",
        "beatport_title", "beatport_artists", "beatport_key", "beatport_key_camelot",
        "beatport_year", "beatport_bpm", "beatport_label", "beatport_genres",
        "beatport_release", "beatport_release_date", "beatport_track_id",
        "beatport_url", "title_sim", "artist_sim", "match_score", "confidence",
        "search_query_index", "search_stop_query_index", "candidate_index"
    ]
    
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for result in results:
            writer.writerow(result.to_dict())
    
    return filepath


def write_candidates_csv(
    results: List[TrackResult],
    base_filename: str,
    output_dir: str = "output"
) -> Optional[str]:
    """
    Write candidates CSV file (all candidates evaluated for all tracks).
    
    Args:
        results: List of TrackResult objects
        base_filename: Base filename with timestamp
        output_dir: Output directory
    
    Returns:
        Path to written file, or None if no candidates
    """
    # Collect all candidates from all tracks
    all_candidates = []
    for result in results:
        all_candidates.extend(result.candidates)
    
    if not all_candidates:
        return None
    
    # Remove .csv extension and add _candidates
    base = base_filename.replace('.csv', '') if base_filename.endswith('.csv') else base_filename
    filepath = os.path.join(output_dir, f"{base}_candidates.csv")
    
    # Get fieldnames from first candidate
    fieldnames = list(all_candidates[0].keys())
    
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_candidates)
    
    return filepath


def write_queries_csv(
    results: List[TrackResult],
    base_filename: str,
    output_dir: str = "output"
) -> Optional[str]:
    """
    Write queries CSV file (all queries executed for all tracks).
    
    Args:
        results: List of TrackResult objects
        base_filename: Base filename with timestamp
        output_dir: Output directory
    
    Returns:
        Path to written file, or None if no queries
    """
    # Collect all queries from all tracks
    all_queries = []
    for result in results:
        all_queries.extend(result.queries)
    
    if not all_queries:
        return None
    
    # Remove .csv extension and add _queries
    base = base_filename.replace('.csv', '') if base_filename.endswith('.csv') else base_filename
    filepath = os.path.join(output_dir, f"{base}_queries.csv")
    
    # Get fieldnames from first query
    fieldnames = list(all_queries[0].keys())
    
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_queries)
    
    return filepath


def write_review_csv(
    results: List[TrackResult],
    review_indices: Set[int],
    base_filename: str,
    output_dir: str = "output"
) -> Optional[str]:
    """
    Write review CSV file (tracks needing manual review).
    
    Args:
        results: List of TrackResult objects
        review_indices: Set of playlist indices that need review
        base_filename: Base filename with timestamp
        output_dir: Output directory
    
    Returns:
        Path to written file, or None if no review tracks
    """
    review_results = [r for r in results if r.playlist_index in review_indices]
    if not review_results:
        return None
    
    # Remove .csv extension and add _review
    base = base_filename.replace('.csv', '') if base_filename.endswith('.csv') else base_filename
    filepath = os.path.join(output_dir, f"{base}_review.csv")
    
    # Use same format as main CSV
    fieldnames = [
        "playlist_index", "original_title", "original_artists",
        "beatport_title", "beatport_artists", "beatport_key", "beatport_key_camelot",
        "beatport_year", "beatport_bpm", "beatport_label", "beatport_genres",
        "beatport_release", "beatport_release_date", "beatport_track_id",
        "beatport_url", "title_sim", "artist_sim", "match_score", "confidence",
        "search_query_index", "search_stop_query_index", "candidate_index"
    ]
    
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for result in review_results:
            writer.writerow(result.to_dict())
    
    return filepath


def write_review_candidates_csv(
    results: List[TrackResult],
    review_indices: Set[int],
    base_filename: str,
    output_dir: str = "output"
) -> Optional[str]:
    """Write candidates CSV for review tracks only"""
    review_results = [r for r in results if r.playlist_index in review_indices]
    if not review_results:
        return None
    
    all_candidates = []
    for result in review_results:
        all_candidates.extend(result.candidates)
    
    if not all_candidates:
        return None
    
    base = base_filename.replace('.csv', '') if base_filename.endswith('.csv') else base_filename
    filepath = os.path.join(output_dir, f"{base}_review_candidates.csv")
    
    fieldnames = list(all_candidates[0].keys())
    
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_candidates)
    
    return filepath


def write_review_queries_csv(
    results: List[TrackResult],
    review_indices: Set[int],
    base_filename: str,
    output_dir: str = "output"
) -> Optional[str]:
    """Write queries CSV for review tracks only"""
    review_results = [r for r in results if r.playlist_index in review_indices]
    if not review_results:
        return None
    
    all_queries = []
    for result in review_results:
        all_queries.extend(result.queries)
    
    if not all_queries:
        return None
    
    base = base_filename.replace('.csv', '') if base_filename.endswith('.csv') else base_filename
    filepath = os.path.join(output_dir, f"{base}_review_queries.csv")
    
    fieldnames = list(all_queries[0].keys())
    
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_queries)
    
    return filepath


def _get_review_indices(results: List[TrackResult]) -> Set[int]:
    """
    Determine which tracks need review based on score and match quality.
    
    Review criteria (from old code):
    - Score < 70
    - Artist similarity < 50 (if artists present)
    - No match found
    """
    review_indices = set()
    
    for result in results:
        needs_review = False
        
        # Check score
        if result.match_score is not None and result.match_score < 70:
            needs_review = True
        
        # Check artist similarity (if artists present)
        if result.artist and result.artist.strip():
            if result.artist_sim is not None and result.artist_sim < 50:
                needs_review = True
        
        # Check if no match
        if not result.matched or not (result.beatport_url or "").strip():
            needs_review = True
        
        if needs_review:
            review_indices.add(result.playlist_index)
    
    return review_indices


def write_json_file(
    results: List[TrackResult],
    file_path: str,
    playlist_name: str = "",
    include_candidates: bool = False,
    include_queries: bool = False
) -> str:
    """
    Write results to JSON file.
    
    Args:
        results: List of TrackResult objects
        file_path: Full path to output JSON file
        playlist_name: Name of playlist (for metadata)
        include_candidates: Whether to include candidates data
        include_queries: Whether to include queries data
    
    Returns:
        Path to written file
    """
    os.makedirs(os.path.dirname(file_path) if os.path.dirname(file_path) else ".", exist_ok=True)
    
    # Build JSON structure
    json_data = {
        "metadata": {
            "playlist_name": playlist_name or "Unknown",
            "processed_at": datetime.now().isoformat(),
            "total_tracks": len(results),
            "matched_tracks": sum(1 for r in results if r.matched),
            "unmatched_tracks": sum(1 for r in results if not r.matched)
        },
        "tracks": []
    }
    
    # Convert tracks to JSON structure
    for result in results:
        track_data = {
            "playlist_index": result.playlist_index,
            "original": {
                "title": result.title,
                "artists": result.artist or ""
            }
        }
        
        if result.matched and result.beatport_url:
            track_data["match"] = {
                "found": True,
                "title": result.beatport_title or "",
                "artists": result.beatport_artists or "",
                "beatport_url": result.beatport_url,
                "beatport_track_id": result.beatport_track_id or "",
                "scores": {
                    "match_score": float(result.match_score) if result.match_score is not None else None,
                    "title_sim": float(result.title_sim) if result.title_sim is not None else None,
                    "artist_sim": float(result.artist_sim) if result.artist_sim is not None else None
                },
                "metadata": {
                    "key": result.beatport_key or "",
                    "key_camelot": result.beatport_key_camelot or "",
                    "year": int(result.beatport_year) if result.beatport_year and result.beatport_year.isdigit() else None,
                    "bpm": int(result.beatport_bpm) if result.beatport_bpm and result.beatport_bpm.isdigit() else None,
                    "label": result.beatport_label or "",
                    "genres": [g.strip() for g in (result.beatport_genres or "").split(",") if g.strip()],
                    "release": result.beatport_release or "",
                    "release_date": result.beatport_release_date or ""
                },
                "confidence": result.confidence or "unknown"
            }
        else:
            track_data["match"] = {
                "found": False
            }
        
        # Add candidates if requested
        if include_candidates and result.candidates:
            track_data["candidates"] = result.candidates
        
        # Add queries if requested
        if include_queries and result.queries:
            track_data["queries"] = result.queries
        
        json_data["tracks"].append(track_data)
    
    # Write JSON file
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(json_data, f, indent=2, ensure_ascii=False)
    
    return file_path


def write_excel_file(
    results: List[TrackResult],
    file_path: str,
    playlist_name: str = ""
) -> str:
    """
    Write results to Excel file (.xlsx).
    
    Args:
        results: List of TrackResult objects
        file_path: Full path to output Excel file
        playlist_name: Name of playlist (for sheet name)
    
    Returns:
        Path to written file
    
    Raises:
        ImportError: If openpyxl is not installed
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install it with: pip install openpyxl"
        )
    
    os.makedirs(os.path.dirname(file_path) if os.path.dirname(file_path) else ".", exist_ok=True)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = playlist_name[:31] if playlist_name else "Results"  # Excel sheet name limit
    
    # Define headers
    headers = [
        "Index", "Title", "Artist", "Matched",
        "Beatport Title", "Beatport Artist", "Score", "Confidence",
        "Key", "BPM", "Year", "Label", "Genres", "Release", "URL"
    ]
    
    # Write headers with styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data rows
    for row_idx, result in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1, value=result.playlist_index)
        ws.cell(row=row_idx, column=2, value=result.title)
        ws.cell(row=row_idx, column=3, value=result.artist or "")
        ws.cell(row=row_idx, column=4, value="Yes" if result.matched else "No")
        ws.cell(row=row_idx, column=5, value=result.beatport_title or "")
        ws.cell(row=row_idx, column=6, value=result.beatport_artists or "")
        ws.cell(row=row_idx, column=7, value=float(result.match_score) if result.match_score is not None else None)
        ws.cell(row=row_idx, column=8, value=result.confidence or "")
        ws.cell(row=row_idx, column=9, value=result.beatport_key_camelot or result.beatport_key or "")
        ws.cell(row=row_idx, column=10, value=int(result.beatport_bpm) if result.beatport_bpm and result.beatport_bpm.isdigit() else None)
        ws.cell(row=row_idx, column=11, value=int(result.beatport_year) if result.beatport_year and result.beatport_year.isdigit() else None)
        ws.cell(row=row_idx, column=12, value=result.beatport_label or "")
        ws.cell(row=row_idx, column=13, value=result.beatport_genres or "")
        ws.cell(row=row_idx, column=14, value=result.beatport_release or "")
        ws.cell(row=row_idx, column=15, value=result.beatport_url or "")
        
        # Color code matched/unmatched rows
        if result.matched:
            fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill
    
    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, start=1):
        column_letter = get_column_letter(col_idx)
        max_length = len(header)
        for row_idx in range(2, len(results) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Save workbook
    wb.save(file_path)
    
    return file_path

